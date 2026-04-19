#!/usr/bin/env python3
"""
CI Efficiency Report Generator

Fetches GitHub PR and workflow data for a list of repositories,
computes CI efficiency metrics, and outputs an Excel report.
"""

from __future__ import annotations

import argparse
import math
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Any

import requests

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


GITHUB_API_BASE = "https://api.github.com"
PER_PAGE = 100
SESSION = requests.Session()
SUMMARY_SHEET = "Management Summary"
APPENDIX_SHEET = "Diagnostic Appendix"
CURRENT_PROBLEMS_SHEET = "Current Problems"
DAILY_DRILLDOWN_SHEET = "Daily Drill-down"
LEGACY_SHEET = "CI效率报告"
WORKFLOW_RAW_SHEET = "Workflow Raw"
JOB_RAW_SHEET = "Job Raw"
STEP_RAW_SHEET = "Step Raw"


@dataclass
class Window:
    start_date: str
    end_date: str

    @property
    def start_ts(self) -> str:
        return f"{self.start_date}T00:00:00Z"

    @property
    def end_ts(self) -> str:
        return f"{self.end_date}T23:59:59Z"

    @property
    def label(self) -> str:
        return f"{self.start_date}..{self.end_date}"


def github_request(token: str, path: str, params: dict | None = None) -> Any:
    headers = {
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    for attempt in range(3):
        try:
            resp = SESSION.get(f"{GITHUB_API_BASE}{path}", headers=headers, params=params, timeout=30)
            if resp.status_code == 404:
                return None
            if resp.status_code == 403 and attempt < 2:
                wait = int(resp.headers.get("Retry-After", "10"))
                print(f"  403, retrying after {wait}s...")
                time.sleep(wait)
                continue
            if resp.status_code == 422:
                print(f"  422: {resp.text[:200]}")
                return None
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.Timeout:
            if attempt < 2:
                time.sleep(2)
                continue
            raise
    return None


def fetch_all_pages_dict(token: str, path: str, params: dict | None = None, items_key: str = "workflow_runs") -> list[dict]:
    all_results: list[dict] = []
    page = 1
    while True:
        page_params = dict(params or {})
        page_params["per_page"] = PER_PAGE
        page_params["page"] = page
        data = github_request(token, path, page_params)
        if not data or not isinstance(data, dict):
            break
        items = data.get(items_key, [])
        if not items:
            break
        all_results.extend(items)
        if len(items) < PER_PAGE:
            break
        page += 1
        time.sleep(0.2)
    return all_results


def fetch_search_results(token: str, query: str, items_key: str = "items") -> list[dict]:
    all_results: list[dict] = []
    page = 1
    while True:
        data = github_request(token, "/search/issues", {
            "q": query,
            "sort": "created",
            "order": "desc",
            "per_page": PER_PAGE,
            "page": page,
        })
        if not data or not isinstance(data, dict):
            break
        items = data.get(items_key, [])
        if not items:
            break
        all_results.extend(items)
        if len(items) < PER_PAGE:
            break
        page += 1
        time.sleep(0.2)
    return all_results


def parse_ts(ts_str: str | None) -> datetime | None:
    if not ts_str:
        return None
    return datetime.fromisoformat(ts_str.replace("Z", "+00:00"))


def diff_minutes(start: str | None, end: str | None) -> float | None:
    start_dt = parse_ts(start)
    end_dt = parse_ts(end)
    if not start_dt or not end_dt:
        return None
    return (end_dt - start_dt).total_seconds() / 60.0


def percentile(values: list[float], p: float) -> float:
    if not values:
        return 0.0
    sorted_vals = sorted(values)
    k = (len(sorted_vals) - 1) * p
    floor = math.floor(k)
    ceil = math.ceil(k)
    if floor == ceil:
        return sorted_vals[int(k)]
    lower = sorted_vals[floor] * (ceil - k)
    upper = sorted_vals[ceil] * (k - floor)
    return lower + upper


def median(values: list[float]) -> float:
    if not values:
        return 0.0
    sorted_vals = sorted(values)
    mid = len(sorted_vals) // 2
    if len(sorted_vals) % 2 == 1:
        return sorted_vals[mid]
    return (sorted_vals[mid - 1] + sorted_vals[mid]) / 2


def average(values: list[float]) -> float:
    return sum(values) / len(values) if values else 0.0


def round1(value: float | None) -> float | None:
    if value is None:
        return None
    return round(value, 1)


def resolve_window(days: int = 90, start_date: str | None = None, end_date: str | None = None) -> Window:
    if start_date or end_date:
        if not start_date or not end_date:
            raise ValueError("Both --start-date and --end-date are required together")
        if start_date > end_date:
            raise ValueError("--start-date must be on or before --end-date")
        return Window(start_date=start_date, end_date=end_date)

    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    return Window(start_date=cutoff.strftime("%Y-%m-%d"), end_date=datetime.now(timezone.utc).strftime("%Y-%m-%d"))


def fetch_merged_prs(token: str, owner: str, repo: str, window: Window) -> list[dict]:
    query = f"repo:{owner}/{repo} is:pr is:merged merged:{window.label}"
    results = fetch_search_results(token, query)
    print(f"  Found {len(results)} merged PRs in window {window.label}")
    return results


def build_bucket_distribution(pr_rows: list[dict]) -> list[dict]:
    ci_rows = [row for row in pr_rows if row.get("ci_e2e_minutes") is not None]
    total = len(ci_rows)
    buckets = [
        ("<60m", lambda value: value < 60),
        ("60-120m", lambda value: 60 <= value < 120),
        ("120-240m", lambda value: 120 <= value < 240),
        (">240m", lambda value: value >= 240),
    ]

    distribution = []
    for label, predicate in buckets:
        matched = [row for row in ci_rows if predicate(row["ci_e2e_minutes"])]
        examples = [
            {
                "number": row["number"],
                "ci_e2e_minutes": row["ci_e2e_minutes"],
                "title": row["title"],
            }
            for row in sorted(matched, key=lambda item: item["ci_e2e_minutes"], reverse=True)[:3]
        ]
        distribution.append({
            "bucket": label,
            "pr_count": len(matched),
            "percentage": round((len(matched) / total) * 100, 1) if total else 0.0,
            "examples": examples,
        })
    return distribution


def classify_month_health(sla_rate: float, severe_long_tail_pct: float) -> str:
    if sla_rate >= 80 and severe_long_tail_pct <= 5:
        return "达标"
    if sla_rate < 50 or severe_long_tail_pct >= 10:
        return "长尾严重"
    return "不达标"


def classify_job_drag(run_count: int, max_runtime: float, avg_runtime: float) -> str:
    if run_count <= 2 and max_runtime >= max(avg_runtime * 1.5, 60):
        return "偶发长尾"
    if run_count >= 3 and avg_runtime >= 30:
        return "高频拖慢项"
    return "需要观察"


def build_anomalies(metrics: dict, distribution: list[dict]) -> list[str]:
    anomalies = []
    severe_bucket = next((bucket for bucket in distribution if bucket["bucket"] == ">240m"), None)

    if metrics["CI E2E达标率(%)"] < 80:
        anomalies.append(f"CI E2E 达标率仅 {metrics['CI E2E达标率(%)']}%，低于 60 分钟目标。")
    if severe_bucket and severe_bucket["percentage"] >= 10:
        anomalies.append(f">240 分钟长尾占比达到 {severe_bucket['percentage']}%，说明长尾问题明显。")
    if metrics["排队耗时 P90(min)"] >= 30:
        anomalies.append(f"排队耗时 P90 为 {metrics['排队耗时 P90(min)']} 分钟，runner 等待已影响体验。")
    if metrics["CI执行时长 P90(min)"] >= 120:
        anomalies.append(f"CI 执行时长 P90 为 {metrics['CI执行时长 P90(min)']} 分钟，主要瓶颈仍在执行阶段。")
    return anomalies


def build_summary_judgment(metrics: dict, distribution: list[dict]) -> str:
    severe_bucket = next((bucket for bucket in distribution if bucket["bucket"] == ">240m"), {"percentage": 0.0})
    status = classify_month_health(metrics["CI E2E达标率(%)"], severe_bucket["percentage"])
    bottleneck = "排队" if metrics["排队耗时 P90(min)"] >= metrics["CI执行时长 P90(min)"] else "执行"
    return (
        f"本月 CI 提交体验{status}，CI E2E 达标率为 {metrics['CI E2E达标率(%)']}%，"
        f"主要矛盾在{bottleneck}阶段。"
    )


def classify_issue_status(run_count: int, drag_type: str) -> str:
    if drag_type == "偶发长尾":
        return "probable_outlier"
    if run_count >= 3 or drag_type == "高频拖慢项":
        return "likely_recurring"
    return "current_issue"


def aggregate_workflows(pr_rows: list[dict]) -> list[dict]:
    grouped: dict[str, dict[str, Any]] = {}

    for pr_row in pr_rows:
        for workflow in pr_row["workflows"]:
            key = workflow["name"]
            entry = grouped.setdefault(key, {
                "workflow_name": key,
                "run_count": 0,
                "total_runtime_minutes": 0.0,
                "max_runtime_minutes": 0.0,
                "queue_samples": [],
                "execution_samples": [],
                "affected_prs": set(),
            })
            runtime = workflow["run_e2e_minutes"] or 0.0
            entry["run_count"] += 1
            entry["total_runtime_minutes"] += runtime
            entry["max_runtime_minutes"] = max(entry["max_runtime_minutes"], runtime)
            entry["affected_prs"].add(pr_row["number"])
            for job in workflow["jobs"]:
                if job["queue_minutes"] is not None:
                    entry["queue_samples"].append(job["queue_minutes"])
                if job["execution_minutes"] is not None:
                    entry["execution_samples"].append(job["execution_minutes"])

    rows = []
    for entry in grouped.values():
        rows.append({
            "workflow_name": entry["workflow_name"],
            "run_count": entry["run_count"],
            "affected_pr_count": len(entry["affected_prs"]),
            "total_runtime_minutes": round(entry["total_runtime_minutes"], 1),
            "avg_runtime_minutes": round(average([entry["total_runtime_minutes"] / entry["run_count"]]), 1),
            "max_runtime_minutes": round(entry["max_runtime_minutes"], 1),
            "queue_p90_minutes": round(percentile(entry["queue_samples"], 0.9), 1) if entry["queue_samples"] else 0.0,
            "execution_p90_minutes": round(percentile(entry["execution_samples"], 0.9), 1) if entry["execution_samples"] else 0.0,
        })
    return sorted(rows, key=lambda row: (row["total_runtime_minutes"], row["max_runtime_minutes"]), reverse=True)


def aggregate_jobs(pr_rows: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str], dict[str, Any]] = {}

    for pr_row in pr_rows:
        for workflow in pr_row["workflows"]:
            for job in workflow["jobs"]:
                key = (workflow["name"], job["name"])
                entry = grouped.setdefault(key, {
                    "workflow_name": workflow["name"],
                    "job_name": job["name"],
                    "run_count": 0,
                    "queue_samples": [],
                    "execution_samples": [],
                })
                entry["run_count"] += 1
                if job["queue_minutes"] is not None:
                    entry["queue_samples"].append(job["queue_minutes"])
                if job["execution_minutes"] is not None:
                    entry["execution_samples"].append(job["execution_minutes"])

    rows = []
    for entry in grouped.values():
        max_runtime = max(entry["execution_samples"], default=0.0)
        avg_runtime = average(entry["execution_samples"])
        total_runtime = sum(entry["execution_samples"])
        rows.append({
            "workflow_name": entry["workflow_name"],
            "job_name": entry["job_name"],
            "run_count": entry["run_count"],
            "max_runtime_minutes": round(max_runtime, 1),
            "avg_runtime_minutes": round(avg_runtime, 1),
            "total_runtime_minutes": round(total_runtime, 1),
            "queue_p90_minutes": round(percentile(entry["queue_samples"], 0.9), 1) if entry["queue_samples"] else 0.0,
            "drag_type": classify_job_drag(entry["run_count"], max_runtime, avg_runtime),
        })
    return sorted(rows, key=lambda row: (row["total_runtime_minutes"], row["max_runtime_minutes"]), reverse=True)


def aggregate_steps(pr_rows: list[dict]) -> tuple[list[dict], bool]:
    grouped: dict[tuple[str, str, str], dict[str, Any]] = {}
    has_partial_coverage = False

    for pr_row in pr_rows:
        for workflow in pr_row["workflows"]:
            for job in workflow["jobs"]:
                if job["steps_partial"]:
                    has_partial_coverage = True
                for step in job["steps"]:
                    key = (workflow["name"], job["name"], step["name"])
                    entry = grouped.setdefault(key, {
                        "workflow_name": workflow["name"],
                        "job_name": job["name"],
                        "step_name": step["name"],
                        "run_count": 0,
                        "samples": [],
                    })
                    entry["run_count"] += 1
                    if step["duration_minutes"] is not None:
                        entry["samples"].append(step["duration_minutes"])

    rows = []
    for entry in grouped.values():
        if not entry["samples"]:
            continue
        rows.append({
            "workflow_name": entry["workflow_name"],
            "job_name": entry["job_name"],
            "step_name": entry["step_name"],
            "run_count": entry["run_count"],
            "max_runtime_minutes": round(max(entry["samples"]), 1),
            "avg_runtime_minutes": round(average(entry["samples"]), 1),
            "total_runtime_minutes": round(sum(entry["samples"]), 1),
        })
    return sorted(rows, key=lambda row: (row["total_runtime_minutes"], row["max_runtime_minutes"]), reverse=True), has_partial_coverage


def build_longest_job_ranking(job_rows: list[dict]) -> list[dict]:
    return sorted(job_rows, key=lambda row: (row["total_runtime_minutes"], row["max_runtime_minutes"]), reverse=True)


def build_daily_problem_list(workflow_rows: list[dict], job_rows: list[dict], step_rows: list[dict], anomalies: list[str]) -> list[dict]:
    problems: list[dict] = []

    for row in workflow_rows[:5]:
        problems.append({
            "Type": "workflow",
            "Name": row["workflow_name"],
            "Scope": f"{row['affected_pr_count']} PRs",
            "Run Count": row["run_count"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Avg Runtime(min)": row["avg_runtime_minutes"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Status": classify_issue_status(row["run_count"], "高频拖慢项" if row["run_count"] >= 3 else "需要观察"),
            "Why It Matters": f"Execution P90 {row['execution_p90_minutes']}m, queue P90 {row['queue_p90_minutes']}m",
        })

    for row in job_rows[:8]:
        problems.append({
            "Type": "job",
            "Name": f"{row['workflow_name']} / {row['job_name']}",
            "Scope": row["workflow_name"],
            "Run Count": row["run_count"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Avg Runtime(min)": row["avg_runtime_minutes"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Status": classify_issue_status(row["run_count"], row["drag_type"]),
            "Why It Matters": row["drag_type"],
        })

    for row in step_rows[:8]:
        problems.append({
            "Type": "step",
            "Name": f"{row['job_name']} / {row['step_name']}",
            "Scope": row["workflow_name"],
            "Run Count": row["run_count"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Avg Runtime(min)": row["avg_runtime_minutes"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Status": classify_issue_status(row["run_count"], "高频拖慢项" if row["run_count"] >= 3 else "需要观察"),
            "Why It Matters": "Step-level hotspot",
        })

    for finding in anomalies:
        problems.append({
            "Type": "signal",
            "Name": finding,
            "Scope": "window",
            "Run Count": "",
            "Max Runtime(min)": "",
            "Avg Runtime(min)": "",
            "Total Runtime(min)": "",
            "Status": "current_issue",
            "Why It Matters": "Window-level anomaly",
        })

    severity_order = {"likely_recurring": 0, "current_issue": 1, "probable_outlier": 2}
    return sorted(
        problems,
        key=lambda row: (
            severity_order.get(str(row["Status"]), 3),
            row["Total Runtime(min)"] if isinstance(row["Total Runtime(min)"], (int, float)) else -1,
            row["Max Runtime(min)"] if isinstance(row["Max Runtime(min)"], (int, float)) else -1,
        ),
        reverse=False,
    )


def build_workflow_raw_rows(report: dict) -> list[dict]:
    rows = []
    for pr_row in report.get("pr_rows", []):
        for workflow in pr_row.get("workflows", []):
            rows.append({
                "repository": report["repo"],
                "pr_number": pr_row.get("number"),
                "pr_title": pr_row.get("title"),
                "pr_url": pr_row.get("html_url"),
                "workflow_run_id": workflow.get("run_id"),
                "workflow_name": workflow.get("name"),
                "workflow_status": workflow.get("status"),
                "workflow_conclusion": workflow.get("conclusion"),
                "branch": workflow.get("branch"),
                "head_sha": workflow.get("head_sha"),
                "event": workflow.get("event"),
                "actor": workflow.get("actor"),
                "created_at": workflow.get("created_at"),
                "started_at": workflow.get("started_at"),
                "completed_at": workflow.get("completed_at"),
                "queue_minutes": workflow.get("queue_minutes"),
                "execution_minutes": workflow.get("execution_minutes"),
                "workflow_url": workflow.get("html_url"),
            })
    return rows


def build_job_raw_rows(report: dict) -> list[dict]:
    rows = []
    for pr_row in report.get("pr_rows", []):
        for workflow in pr_row.get("workflows", []):
            for job in workflow.get("jobs", []):
                rows.append({
                    "repository": report["repo"],
                    "pr_number": pr_row.get("number"),
                    "workflow_run_id": workflow.get("run_id"),
                    "workflow_name": workflow.get("name"),
                    "job_id": job.get("job_id"),
                    "job_name": job.get("name"),
                    "runner_name": job.get("runner_name"),
                    "runner_group": job.get("runner_group"),
                    "runner_labels": ", ".join(job.get("runner_labels", [])),
                    "job_status": job.get("status"),
                    "job_conclusion": job.get("conclusion"),
                    "created_at": job.get("created_at"),
                    "started_at": job.get("started_at"),
                    "completed_at": job.get("completed_at"),
                    "queue_minutes": job.get("queue_minutes"),
                    "execution_minutes": job.get("execution_minutes"),
                    "html_url": job.get("html_url"),
                })
    return rows


def build_step_raw_rows(report: dict) -> list[dict]:
    rows = []
    for pr_row in report.get("pr_rows", []):
        for workflow in pr_row.get("workflows", []):
            for job in workflow.get("jobs", []):
                for step in job.get("steps", []):
                    rows.append({
                        "repository": report["repo"],
                        "pr_number": pr_row.get("number"),
                        "workflow_run_id": workflow.get("run_id"),
                        "workflow_name": workflow.get("name"),
                        "job_id": job.get("job_id"),
                        "job_name": job.get("name"),
                        "step_number": step.get("number"),
                        "step_name": step.get("name"),
                        "step_status": step.get("status"),
                        "step_conclusion": step.get("conclusion"),
                        "started_at": step.get("started_at"),
                        "completed_at": step.get("completed_at"),
                        "execution_minutes": step.get("duration_minutes"),
                        "raw_step_index": step.get("raw_step_index"),
                    })
    return rows


def build_management_metrics(repo: str, window: Window, pr_rows: list[dict], review_reliable: bool) -> dict:
    pr_e2e = [row["pr_e2e_minutes"] for row in pr_rows if row["pr_e2e_minutes"] is not None]
    ci_e2e = [row["ci_e2e_minutes"] for row in pr_rows if row["ci_e2e_minutes"] is not None]
    queue = [row["max_queue_minutes"] for row in pr_rows if row["max_queue_minutes"] is not None]
    execution = [row["max_execution_minutes"] for row in pr_rows if row["max_execution_minutes"] is not None]
    review = [row["review_minutes"] for row in pr_rows if row["review_minutes"] is not None]

    metrics = {
        "Repository": repo,
        "WindowStart": window.start_ts,
        "WindowEnd": window.end_ts,
        "PR E2E时长 P90(min)": round(percentile(pr_e2e, 0.9), 1),
        "CI E2E时长 P50(min)": round(median(ci_e2e), 1),
        "CI E2E时长 P90(min)": round(percentile(ci_e2e, 0.9), 1),
        "排队耗时 P90(min)": round(percentile(queue, 0.9), 1),
        "CI执行时长 P90(min)": round(percentile(execution, 0.9), 1),
        "PR检视时长 P90(min)": round(percentile(review, 0.9), 1) if review_reliable else None,
        "CI E2E达标率(%)": round(sum(1 for value in ci_e2e if value < 60) / len(ci_e2e) * 100, 1) if ci_e2e else 0.0,
        "统计PR数": len(pr_rows),
    }
    return metrics


def compute_repo_report(token: str, owner: str, repo: str, window: Window, max_prs: int = 0) -> dict:
    full_repo = f"{owner}/{repo}"
    print(f"\nProcessing {full_repo}...")

    prs = fetch_merged_prs(token, owner, repo, window)
    if max_prs > 0:
        prs = prs[:max_prs]
    if not prs:
        empty_metrics = {
            "Repository": full_repo,
            "WindowStart": window.start_ts,
            "WindowEnd": window.end_ts,
            "PR E2E时长 P90(min)": 0.0,
            "CI E2E时长 P50(min)": 0.0,
            "CI E2E时长 P90(min)": 0.0,
            "排队耗时 P90(min)": 0.0,
            "CI执行时长 P90(min)": 0.0,
            "PR检视时长 P90(min)": None,
            "CI E2E达标率(%)": 0.0,
            "统计PR数": 0,
        }
        return {
            "repo": full_repo,
            "window": window,
            "metrics": empty_metrics,
            "pr_rows": [],
            "distribution": [],
            "workflow_rows": [],
            "job_rows": [],
            "step_rows": [],
            "longest_jobs": [],
            "anomalies": ["No merged PRs found in the selected window."],
            "summary_judgment": "所选时间窗内没有 merged PR，无法评估月度 CI 提交体验。",
            "coverage_notes": ["No merged PRs found in the selected window."],
            "legacy_row": empty_metrics,
        }

    print(f"  Fetching PR details (head_sha)...")
    pr_list = []
    for index, pr in enumerate(prs, 1):
        if index % 50 == 0 or index == len(prs):
            print(f"  Fetching PR details {index}/{len(prs)}...")
        detail = github_request(token, f"/repos/{owner}/{repo}/pulls/{pr['number']}")
        if not detail:
            continue
        pr_list.append({
            "number": pr["number"],
            "title": detail["title"],
            "created_at": detail["created_at"],
            "merged_at": detail.get("merged_at") or pr.get("closed_at"),
            "head_sha": detail.get("head", {}).get("sha"),
            "html_url": detail["html_url"],
        })
        time.sleep(0.2)

    pr_rows = []
    review_samples = 0
    review_reliable = True

    for idx, pr_info in enumerate(pr_list, 1):
        head_sha = pr_info["head_sha"]
        runs = fetch_all_pages_dict(token, f"/repos/{owner}/{repo}/actions/runs", {
            "head_sha": head_sha,
            "status": "completed",
            "created": window.label,
        }) if head_sha else []

        workflows = []
        latest_ci_complete = None
        max_queue = None
        max_execution = None

        for run in runs:
            jobs = fetch_all_pages_dict(token, f"/repos/{owner}/{repo}/actions/runs/{run['id']}/jobs", {"filter": "latest"}, "jobs")
            job_rows = []
            workflow_queue_samples = []
            workflow_execution_samples = []

            for job in jobs:
                queue_minutes = diff_minutes(job.get("created_at"), job.get("started_at"))
                execution_minutes = diff_minutes(job.get("started_at"), job.get("completed_at"))
                if queue_minutes is not None:
                    max_queue = queue_minutes if max_queue is None else max(max_queue, queue_minutes)
                    workflow_queue_samples.append(queue_minutes)
                if execution_minutes is not None:
                    max_execution = execution_minutes if max_execution is None else max(max_execution, execution_minutes)
                    workflow_execution_samples.append(execution_minutes)

                steps_partial = False
                steps = []
                for step_index, step in enumerate(job.get("steps", []) or [], start=1):
                    duration_minutes = diff_minutes(step.get("started_at"), step.get("completed_at"))
                    if duration_minutes is None:
                        steps_partial = True
                    steps.append({
                        "number": step.get("number"),
                        "name": step.get("name") or f"step-{step.get('number', 'unknown')}",
                        "status": step.get("status"),
                        "conclusion": step.get("conclusion"),
                        "started_at": step.get("started_at"),
                        "completed_at": step.get("completed_at"),
                        "duration_minutes": round1(duration_minutes),
                        "raw_step_index": step_index,
                    })

                job_rows.append({
                    "job_id": job.get("id"),
                    "name": job.get("name", f"job-{job.get('id')}"),
                    "workflow_name": job.get("workflow_name") or run.get("name") or "Unknown Workflow",
                    "runner_name": job.get("runner_name"),
                    "runner_group": job.get("runner_group_name"),
                    "runner_labels": job.get("labels") or [],
                    "status": job.get("status"),
                    "conclusion": job.get("conclusion"),
                    "created_at": job.get("created_at"),
                    "started_at": job.get("started_at"),
                    "completed_at": job.get("completed_at"),
                    "queue_minutes": round1(queue_minutes),
                    "execution_minutes": round1(execution_minutes),
                    "html_url": job.get("html_url"),
                    "steps": steps,
                    "steps_partial": steps_partial,
                })

            run_complete_ts = run.get("updated_at") or run.get("completed_at")
            run_complete = parse_ts(run_complete_ts)
            if run_complete and (latest_ci_complete is None or run_complete > latest_ci_complete):
                latest_ci_complete = run_complete

            workflow_name = run.get("name") or next((job["workflow_name"] for job in job_rows if job["workflow_name"]), "Unknown Workflow")
            workflows.append({
                "run_id": run.get("id"),
                "name": workflow_name,
                "status": run.get("status"),
                "conclusion": run.get("conclusion"),
                "branch": run.get("head_branch"),
                "head_sha": run.get("head_sha"),
                "event": run.get("event"),
                "actor": (run.get("actor") or {}).get("login"),
                "created_at": run.get("created_at"),
                "started_at": run.get("run_started_at") or run.get("started_at"),
                "completed_at": run_complete_ts,
                "queue_minutes": round1(percentile(workflow_queue_samples, 0.9)) if workflow_queue_samples else None,
                "execution_minutes": round1(percentile(workflow_execution_samples, 0.9)) if workflow_execution_samples else None,
                "run_e2e_minutes": round1(diff_minutes(run.get("created_at"), run_complete_ts)),
                "html_url": run.get("html_url"),
                "jobs": job_rows,
            })

        pr_e2e_minutes = round1(diff_minutes(pr_info["created_at"], pr_info["merged_at"]))
        ci_e2e_minutes = round1(max((workflow["run_e2e_minutes"] or 0 for workflow in workflows), default=0.0)) if workflows else None
        review_minutes = None
        if latest_ci_complete and pr_info["merged_at"]:
            merged_dt = parse_ts(pr_info["merged_at"])
            review_delta = (merged_dt - latest_ci_complete).total_seconds() / 60.0 if merged_dt else None
            if review_delta is not None and review_delta >= 0:
                review_minutes = round(review_delta, 1)
                review_samples += 1

        if workflows and latest_ci_complete is None:
            review_reliable = False

        pr_rows.append({
            "number": pr_info["number"],
            "title": pr_info["title"],
            "html_url": pr_info["html_url"],
            "created_at": pr_info["created_at"],
            "merged_at": pr_info["merged_at"],
            "pr_e2e_minutes": pr_e2e_minutes,
            "ci_e2e_minutes": ci_e2e_minutes,
            "review_minutes": review_minutes,
            "max_queue_minutes": round1(max_queue),
            "max_execution_minutes": round1(max_execution),
            "workflows": workflows,
        })

        if idx % 10 == 0 or idx == len(pr_list):
            print(f"  Processed {idx}/{len(pr_list)} PRs")

    if review_samples == 0:
        review_reliable = False

    workflow_rows = aggregate_workflows(pr_rows)
    job_rows = aggregate_jobs(pr_rows)
    step_rows, has_partial_step_coverage = aggregate_steps(pr_rows)
    longest_jobs = build_longest_job_ranking(job_rows)[:20]
    distribution = build_bucket_distribution(pr_rows)
    metrics = build_management_metrics(full_repo, window, pr_rows, review_reliable)
    anomalies = build_anomalies(metrics, distribution)
    summary_judgment = build_summary_judgment(metrics, distribution)
    daily_problems = build_daily_problem_list(workflow_rows, job_rows, step_rows, anomalies)
    coverage_notes = []
    if not review_reliable:
        coverage_notes.append("PR review metrics are not reliable enough for Management Summary in this window.")
    if has_partial_step_coverage:
        coverage_notes.append("Step-level timing coverage is partial; step analysis should be treated as diagnostic appendix only.")

    legacy_row = {
        "Repository": metrics["Repository"],
        "PR E2E时长 P90(min)": metrics["PR E2E时长 P90(min)"],
        "CI E2E时长 P90(min)": metrics["CI E2E时长 P90(min)"],
        "排队耗时 P90(min)": metrics["排队耗时 P90(min)"],
        "CI执行时长 P90(min)": metrics["CI执行时长 P90(min)"],
        "PR检视时长 P90(min)": metrics["PR检视时长 P90(min)"] if metrics["PR检视时长 P90(min)"] is not None else "Not reliable",
        "CI E2E达标率(%)": metrics["CI E2E达标率(%)"],
        "统计PR数": metrics["统计PR数"],
    }

    return {
        "repo": full_repo,
        "window": window,
        "metrics": metrics,
        "pr_rows": pr_rows,
        "distribution": distribution,
        "workflow_rows": workflow_rows,
        "job_rows": job_rows,
        "step_rows": step_rows,
        "longest_jobs": longest_jobs,
        "anomalies": anomalies,
        "summary_judgment": summary_judgment,
        "daily_problems": daily_problems,
        "coverage_notes": coverage_notes,
        "legacy_row": legacy_row,
    }


def set_header(cell) -> None:
    cell.font = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def set_cell(cell, value: Any) -> None:
    cell.value = value
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def autofit_columns(ws) -> None:
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 36)


def write_table(ws, start_row: int, title: str, columns: list[str], rows: list[dict]) -> int:
    ws.cell(row=start_row, column=1, value=title).font = Font(name="Microsoft YaHei", bold=True, size=12)
    header_row = start_row + 1
    for idx, column in enumerate(columns, 1):
        set_header(ws.cell(row=header_row, column=idx))
        ws.cell(row=header_row, column=idx, value=column)

    row_index = header_row + 1
    for row in rows:
        for idx, column in enumerate(columns, 1):
            set_cell(ws.cell(row=row_index, column=idx), row.get(column, ""))
        row_index += 1
    return row_index + 1


def write_legacy_summary_sheet(wb: openpyxl.Workbook, reports: list[dict]) -> None:
    ws = wb.active
    ws.title = LEGACY_SHEET
    columns = [
        "Repository",
        "PR E2E时长 P90(min)",
        "CI E2E时长 P90(min)",
        "排队耗时 P90(min)",
        "CI执行时长 P90(min)",
        "PR检视时长 P90(min)",
        "CI E2E达标率(%)",
        "统计PR数",
    ]
    for idx, column in enumerate(columns, 1):
        set_header(ws.cell(row=1, column=idx))
        ws.cell(row=1, column=idx, value=column)
    for row_index, report in enumerate(reports, 2):
        for col_index, column in enumerate(columns, 1):
            value = report["legacy_row"].get(column, "")
            set_cell(ws.cell(row=row_index, column=col_index), value)
            if column == "CI E2E达标率(%)" and isinstance(value, (int, float)):
                fill = "C6EFCE" if value >= 80 else "FFEB9C" if value >= 50 else "FFC7CE"
                ws.cell(row=row_index, column=col_index).fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
    ws.freeze_panes = "A2"
    autofit_columns(ws)


def write_management_summary_sheet(wb: openpyxl.Workbook, report: dict) -> None:
    ws = wb.create_sheet(title=SUMMARY_SHEET)
    metrics = report["metrics"]

    ws["A1"] = "Monthly CI Experience Report"
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14)
    ws["A2"] = f"Repository: {report['repo']}"
    ws["A3"] = f"Window: {report['window'].start_date} to {report['window'].end_date}"
    ws["A4"] = report["summary_judgment"]

    summary_rows = [
        {"Metric": "统计PR数", "Value": metrics["统计PR数"]},
        {"Metric": "CI E2E P50(min)", "Value": metrics["CI E2E时长 P50(min)"]},
        {"Metric": "CI E2E P90(min)", "Value": metrics["CI E2E时长 P90(min)"]},
        {"Metric": "排队耗时 P90(min)", "Value": metrics["排队耗时 P90(min)"]},
        {"Metric": "CI执行时长 P90(min)", "Value": metrics["CI执行时长 P90(min)"]},
        {"Metric": "CI E2E达标率(%)", "Value": metrics["CI E2E达标率(%)"]},
    ]
    next_row = write_table(ws, 6, "Core Metrics", ["Metric", "Value"], summary_rows)

    distribution_rows = []
    for bucket in report["distribution"]:
        examples = ", ".join(f"#{item['number']} ({item['ci_e2e_minutes']}m)" for item in bucket["examples"])
        distribution_rows.append({
            "Bucket": bucket["bucket"],
            "PR Count": bucket["pr_count"],
            "Percentage": bucket["percentage"],
            "Examples": examples,
        })
    next_row = write_table(ws, next_row, "CI E2E Distribution", ["Bucket", "PR Count", "Percentage", "Examples"], distribution_rows)

    anomaly_rows = [{"Finding": finding} for finding in (report["anomalies"] or ["No notable anomalies detected."])]
    next_row = write_table(ws, next_row, "Anomalies", ["Finding"], anomaly_rows)

    top_workflows = [
        {
            "Workflow": row["workflow_name"],
            "Run Count": row["run_count"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Execution P90(min)": row["execution_p90_minutes"],
        }
        for row in report["workflow_rows"][:10]
    ]
    next_row = write_table(ws, next_row, "Top Workflows", ["Workflow", "Run Count", "Total Runtime(min)", "Execution P90(min)"], top_workflows)

    top_jobs = [
        {
            "Workflow": row["workflow_name"],
            "Job": row["job_name"],
            "Run Count": row["run_count"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Drag Type": row["drag_type"],
        }
        for row in report["longest_jobs"][:10]
    ]
    next_row = write_table(ws, next_row, "Top Jobs", ["Workflow", "Job", "Run Count", "Total Runtime(min)", "Drag Type"], top_jobs)

    if report["coverage_notes"]:
        coverage_rows = [{"Note": note} for note in report["coverage_notes"]]
        write_table(ws, next_row, "Coverage Notes", ["Note"], coverage_rows)

    ws.freeze_panes = "A7"
    autofit_columns(ws)


def write_diagnostic_appendix_sheet(wb: openpyxl.Workbook, report: dict) -> None:
    ws = wb.create_sheet(title=APPENDIX_SHEET)
    ws["A1"] = "Diagnostic Appendix"
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14)
    ws["A2"] = f"Repository: {report['repo']}"
    ws["A3"] = f"Window: {report['window'].start_date} to {report['window'].end_date}"

    next_row = 5
    workflow_rows = [
        {
            "Workflow": row["workflow_name"],
            "Run Count": row["run_count"],
            "Affected PRs": row["affected_pr_count"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Avg Runtime(min)": row["avg_runtime_minutes"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Queue P90(min)": row["queue_p90_minutes"],
            "Execution P90(min)": row["execution_p90_minutes"],
        }
        for row in report["workflow_rows"][:50]
    ]
    next_row = write_table(
        ws,
        next_row,
        "Workflow Ranking",
        ["Workflow", "Run Count", "Affected PRs", "Total Runtime(min)", "Avg Runtime(min)", "Max Runtime(min)", "Queue P90(min)", "Execution P90(min)"],
        workflow_rows,
    )

    job_rows = [
        {
            "Workflow": row["workflow_name"],
            "Job": row["job_name"],
            "Run Count": row["run_count"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Avg Runtime(min)": row["avg_runtime_minutes"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Queue P90(min)": row["queue_p90_minutes"],
            "Drag Type": row["drag_type"],
        }
        for row in report["job_rows"][:100]
    ]
    next_row = write_table(
        ws,
        next_row,
        "Job Ranking",
        ["Workflow", "Job", "Run Count", "Max Runtime(min)", "Avg Runtime(min)", "Total Runtime(min)", "Queue P90(min)", "Drag Type"],
        job_rows,
    )

    step_rows = [
        {
            "Workflow": row["workflow_name"],
            "Job": row["job_name"],
            "Step": row["step_name"],
            "Run Count": row["run_count"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Avg Runtime(min)": row["avg_runtime_minutes"],
            "Total Runtime(min)": row["total_runtime_minutes"],
        }
        for row in report["step_rows"][:100]
    ]
    next_row = write_table(
        ws,
        next_row,
        "Step Ranking",
        ["Workflow", "Job", "Step", "Run Count", "Max Runtime(min)", "Avg Runtime(min)", "Total Runtime(min)"],
        step_rows or [{"Workflow": "", "Job": "", "Step": "No step-level timing rows available", "Run Count": "", "Max Runtime(min)": "", "Avg Runtime(min)": "", "Total Runtime(min)": ""}],
    )

    longest_job_rows = [
        {
            "Workflow": row["workflow_name"],
            "Job": row["job_name"],
            "Run Count": row["run_count"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Avg Runtime(min)": row["avg_runtime_minutes"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Drag Type": row["drag_type"],
        }
        for row in report["longest_jobs"][:50]
    ]
    next_row = write_table(
        ws,
        next_row,
        "Longest Job Ranking",
        ["Workflow", "Job", "Run Count", "Max Runtime(min)", "Avg Runtime(min)", "Total Runtime(min)", "Drag Type"],
        longest_job_rows,
    )

    if report["coverage_notes"]:
        coverage_rows = [{"Note": note} for note in report["coverage_notes"]]
        write_table(ws, next_row, "Coverage Notes", ["Note"], coverage_rows)

    ws.freeze_panes = "A6"
    autofit_columns(ws)


def write_daily_current_problems_sheet(wb: openpyxl.Workbook, report: dict) -> None:
    ws = wb.create_sheet(title=CURRENT_PROBLEMS_SHEET)
    ws["A1"] = "Daily CI Diagnostic"
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14)
    ws["A2"] = f"Repository: {report['repo']}"
    ws["A3"] = f"Window: {report['window'].start_date} to {report['window'].end_date}"
    ws["A4"] = (
        "Current issues detected in the selected window."
        if report["daily_problems"]
        else "No major CI problems detected in the selected window."
    )

    problem_rows = report["daily_problems"][:20] or [{
        "Type": "signal",
        "Name": "No major issues detected",
        "Scope": "window",
        "Run Count": "",
        "Max Runtime(min)": "",
        "Avg Runtime(min)": "",
        "Total Runtime(min)": "",
        "Status": "current_issue",
        "Why It Matters": "Nothing exceeded the current problem thresholds.",
    }]
    next_row = write_table(
        ws,
        6,
        "Current Problem List",
        ["Type", "Name", "Scope", "Run Count", "Max Runtime(min)", "Avg Runtime(min)", "Total Runtime(min)", "Status", "Why It Matters"],
        problem_rows,
    )

    if report["coverage_notes"]:
        coverage_rows = [{"Note": note} for note in report["coverage_notes"]]
        write_table(ws, next_row, "Coverage Notes", ["Note"], coverage_rows)

    ws.freeze_panes = "A7"
    autofit_columns(ws)


def write_daily_drilldown_sheet(wb: openpyxl.Workbook, report: dict) -> None:
    ws = wb.create_sheet(title=DAILY_DRILLDOWN_SHEET)
    ws["A1"] = "Daily Drill-down"
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14)
    ws["A2"] = f"Repository: {report['repo']}"
    ws["A3"] = f"Window: {report['window'].start_date} to {report['window'].end_date}"

    next_row = 5
    problem_rows = report["daily_problems"][:20]
    if problem_rows:
        next_row = write_table(
            ws,
            next_row,
            "Problem Ranking",
            ["Type", "Name", "Scope", "Run Count", "Max Runtime(min)", "Avg Runtime(min)", "Total Runtime(min)", "Status", "Why It Matters"],
            problem_rows,
        )

    workflow_rows = [
        {
            "Workflow": row["workflow_name"],
            "Run Count": row["run_count"],
            "Affected PRs": row["affected_pr_count"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Execution P90(min)": row["execution_p90_minutes"],
        }
        for row in report["workflow_rows"][:30]
    ]
    next_row = write_table(
        ws,
        next_row,
        "Workflow Hotspots",
        ["Workflow", "Run Count", "Affected PRs", "Total Runtime(min)", "Max Runtime(min)", "Execution P90(min)"],
        workflow_rows,
    )

    job_rows = [
        {
            "Workflow": row["workflow_name"],
            "Job": row["job_name"],
            "Run Count": row["run_count"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Avg Runtime(min)": row["avg_runtime_minutes"],
            "Total Runtime(min)": row["total_runtime_minutes"],
            "Drag Type": row["drag_type"],
        }
        for row in report["job_rows"][:50]
    ]
    next_row = write_table(
        ws,
        next_row,
        "Job Hotspots",
        ["Workflow", "Job", "Run Count", "Max Runtime(min)", "Avg Runtime(min)", "Total Runtime(min)", "Drag Type"],
        job_rows,
    )

    step_rows = [
        {
            "Workflow": row["workflow_name"],
            "Job": row["job_name"],
            "Step": row["step_name"],
            "Run Count": row["run_count"],
            "Max Runtime(min)": row["max_runtime_minutes"],
            "Avg Runtime(min)": row["avg_runtime_minutes"],
            "Total Runtime(min)": row["total_runtime_minutes"],
        }
        for row in report["step_rows"][:50]
    ]
    next_row = write_table(
        ws,
        next_row,
        "Step Hotspots",
        ["Workflow", "Job", "Step", "Run Count", "Max Runtime(min)", "Avg Runtime(min)", "Total Runtime(min)"],
        step_rows or [{"Workflow": "", "Job": "", "Step": "No step-level timing rows available", "Run Count": "", "Max Runtime(min)": "", "Avg Runtime(min)": "", "Total Runtime(min)": ""}],
    )

    if report["coverage_notes"]:
        coverage_rows = [{"Note": note} for note in report["coverage_notes"]]
        write_table(ws, next_row, "Coverage Notes", ["Note"], coverage_rows)

    ws.freeze_panes = "A6"
    autofit_columns(ws)


def write_raw_data_sheet(wb: openpyxl.Workbook, title: str, columns: list[str], rows: list[dict], empty_message: str) -> None:
    ws = wb.create_sheet(title=title)
    for idx, column in enumerate(columns, 1):
        set_header(ws.cell(row=1, column=idx))
        ws.cell(row=1, column=idx, value=column)

    if rows:
        for row_index, row in enumerate(rows, start=2):
            for col_index, column in enumerate(columns, start=1):
                set_cell(ws.cell(row=row_index, column=col_index), row.get(column, ""))
    else:
        set_cell(ws.cell(row=2, column=1), empty_message)

    ws.freeze_panes = "A2"
    autofit_columns(ws)


def write_raw_data_sheets(wb: openpyxl.Workbook, report: dict) -> None:
    workflow_columns = [
        "repository",
        "pr_number",
        "pr_title",
        "pr_url",
        "workflow_run_id",
        "workflow_name",
        "workflow_status",
        "workflow_conclusion",
        "branch",
        "head_sha",
        "event",
        "actor",
        "created_at",
        "started_at",
        "completed_at",
        "queue_minutes",
        "execution_minutes",
        "workflow_url",
    ]
    job_columns = [
        "repository",
        "pr_number",
        "workflow_run_id",
        "workflow_name",
        "job_id",
        "job_name",
        "runner_name",
        "runner_group",
        "runner_labels",
        "job_status",
        "job_conclusion",
        "created_at",
        "started_at",
        "completed_at",
        "queue_minutes",
        "execution_minutes",
        "html_url",
    ]
    step_columns = [
        "repository",
        "pr_number",
        "workflow_run_id",
        "workflow_name",
        "job_id",
        "job_name",
        "step_number",
        "step_name",
        "step_status",
        "step_conclusion",
        "started_at",
        "completed_at",
        "execution_minutes",
        "raw_step_index",
    ]

    write_raw_data_sheet(
        wb,
        WORKFLOW_RAW_SHEET,
        workflow_columns,
        build_workflow_raw_rows(report),
        "No workflow raw rows available",
    )
    write_raw_data_sheet(
        wb,
        JOB_RAW_SHEET,
        job_columns,
        build_job_raw_rows(report),
        "No job raw rows available",
    )
    write_raw_data_sheet(
        wb,
        STEP_RAW_SHEET,
        step_columns,
        build_step_raw_rows(report),
        "No step raw rows available",
    )


def write_excel(reports: list[dict], output_path: str, report_mode: str) -> None:
    wb = openpyxl.Workbook()
    write_legacy_summary_sheet(wb, reports)

    if len(reports) == 1:
        if report_mode == "monthly_summary":
            write_management_summary_sheet(wb, reports[0])
            write_diagnostic_appendix_sheet(wb, reports[0])
        elif report_mode == "daily_diagnostic":
            write_daily_current_problems_sheet(wb, reports[0])
            write_daily_drilldown_sheet(wb, reports[0])
        else:
            raise ValueError(f"Unsupported report mode: {report_mode}")
        write_raw_data_sheets(wb, reports[0])

    wb.save(output_path)
    print(f"\nReport saved to: {output_path}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate CI efficiency report for GitHub repositories")
    parser.add_argument("--repos", nargs="+", required=True, help="List of org/repo")
    parser.add_argument("--token", required=True, help="GitHub PAT")
    parser.add_argument("--output", default="ci_efficiency_report.xlsx", help="Output Excel file path")
    parser.add_argument("--report-mode", default="monthly_summary", choices=["monthly_summary", "daily_diagnostic"], help="Report mode")
    parser.add_argument("--days", type=int, default=90, help="Number of days to look back when start/end are not provided")
    parser.add_argument("--start-date", help="Start date in YYYY-MM-DD format")
    parser.add_argument("--end-date", help="End date in YYYY-MM-DD format")
    parser.add_argument("--max-prs", type=int, default=0, help="Max PRs per repo (0=unlimited)")
    args = parser.parse_args()

    try:
        window = resolve_window(days=args.days, start_date=args.start_date, end_date=args.end_date)
    except ValueError as exc:
        print(f"Error: {exc}")
        sys.exit(1)

    reports = []
    for repo_str in args.repos:
        if "/" not in repo_str:
            print(f"Skipping invalid repo format: {repo_str} (expected org/repo)")
            continue
        owner, repo = repo_str.split("/", 1)
        report = compute_repo_report(args.token, owner, repo, window, args.max_prs)
        reports.append(report)

    if not reports:
        print("No valid repositories to process.")
        sys.exit(1)

    write_excel(reports, args.output, args.report_mode)


if __name__ == "__main__":
    main()
