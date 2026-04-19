import importlib.util
import sys
import unittest
from pathlib import Path
from tempfile import NamedTemporaryFile

from openpyxl import load_workbook


SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "ci_efficiency_report.py"
SPEC = importlib.util.spec_from_file_location("ci_efficiency_report", SCRIPT_PATH)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC.loader is not None
sys.modules[SPEC.name] = MODULE
SPEC.loader.exec_module(MODULE)


class CiEfficiencyReportTests(unittest.TestCase):
    def test_resolve_window_with_explicit_dates(self):
        window = MODULE.resolve_window(start_date="2026-04-01", end_date="2026-04-30")
        self.assertEqual(window.start_date, "2026-04-01")
        self.assertEqual(window.end_date, "2026-04-30")
        self.assertEqual(window.label, "2026-04-01..2026-04-30")

    def test_build_bucket_distribution_uses_four_fixed_buckets(self):
        pr_rows = [
            {"number": 1, "title": "a", "ci_e2e_minutes": 30},
            {"number": 2, "title": "b", "ci_e2e_minutes": 90},
            {"number": 3, "title": "c", "ci_e2e_minutes": 180},
            {"number": 4, "title": "d", "ci_e2e_minutes": 360},
        ]

        buckets = MODULE.build_bucket_distribution(pr_rows)

        self.assertEqual([bucket["bucket"] for bucket in buckets], ["<60m", "60-120m", "120-240m", ">240m"])
        self.assertEqual([bucket["pr_count"] for bucket in buckets], [1, 1, 1, 1])
        self.assertEqual([bucket["percentage"] for bucket in buckets], [25.0, 25.0, 25.0, 25.0])

    def test_workflow_job_step_aggregation_and_drag_classification(self):
        pr_rows = [
            {
                "number": 101,
                "title": "First",
                "ci_e2e_minutes": 220.0,
                "workflows": [
                    {
                        "name": "E2E-Full",
                        "run_e2e_minutes": 180.0,
                        "jobs": [
                            {
                                "name": "integration-tests",
                                "queue_minutes": 10.0,
                                "execution_minutes": 120.0,
                                "steps": [
                                    {"name": "setup", "duration_minutes": 20.0},
                                    {"name": "run tests", "duration_minutes": 80.0},
                                ],
                                "steps_partial": False,
                            },
                            {
                                "name": "report",
                                "queue_minutes": 2.0,
                                "execution_minutes": 15.0,
                                "steps": [
                                    {"name": "upload", "duration_minutes": 8.0},
                                ],
                                "steps_partial": False,
                            },
                        ],
                    }
                ],
            },
            {
                "number": 102,
                "title": "Second",
                "ci_e2e_minutes": 240.0,
                "workflows": [
                    {
                        "name": "E2E-Full",
                        "run_e2e_minutes": 200.0,
                        "jobs": [
                            {
                                "name": "integration-tests",
                                "queue_minutes": 12.0,
                                "execution_minutes": 130.0,
                                "steps": [
                                    {"name": "setup", "duration_minutes": 18.0},
                                    {"name": "run tests", "duration_minutes": 90.0},
                                ],
                                "steps_partial": False,
                            }
                        ],
                    }
                ],
            },
            {
                "number": 103,
                "title": "Third",
                "ci_e2e_minutes": 75.0,
                "workflows": [
                    {
                        "name": "Lint",
                        "run_e2e_minutes": 20.0,
                        "jobs": [
                            {
                                "name": "lint",
                                "queue_minutes": 1.0,
                                "execution_minutes": 65.0,
                                "steps": [
                                    {"name": "run lint", "duration_minutes": 60.0},
                                ],
                                "steps_partial": False,
                            }
                        ],
                    }
                ],
            },
        ]

        report = {"repo": "owner/repo", "pr_rows": pr_rows}
        workflow_raw_rows = MODULE.build_workflow_raw_rows(report)
        job_raw_rows = MODULE.build_job_raw_rows(report)
        step_raw_rows = MODULE.build_step_raw_rows(report)
        workflow_rows = MODULE.aggregate_workflows_from_raw_rows(workflow_raw_rows)
        job_rows = MODULE.aggregate_jobs_from_raw_rows(job_raw_rows)
        step_rows, has_partial = MODULE.aggregate_steps_from_raw_rows(step_raw_rows)
        longest_jobs = MODULE.build_longest_job_ranking(job_rows)

        self.assertFalse(has_partial)
        self.assertEqual(workflow_rows[0]["workflow_name"], "E2E-Full")
        self.assertEqual(workflow_rows[0]["run_count"], 2)
        self.assertEqual(job_rows[0]["job_name"], "integration-tests")
        self.assertEqual(job_rows[0]["run_count"], 2)
        self.assertEqual(job_rows[0]["drag_type"], "需要观察")
        self.assertEqual(step_rows[0]["step_name"], "run tests")
        self.assertEqual(longest_jobs[0]["job_name"], "integration-tests")

    def test_longest_job_marks_rare_outlier(self):
        job_rows = [
            {
                "workflow_name": "Nightly",
                "job_name": "huge-job",
                "run_count": 1,
                "max_runtime_minutes": 200.0,
                "avg_runtime_minutes": 100.0,
                "total_runtime_minutes": 100.0,
                "queue_p90_minutes": 0.0,
                "drag_type": MODULE.classify_job_drag(1, 200.0, 100.0),
            }
        ]

        ranking = MODULE.build_longest_job_ranking(job_rows)
        self.assertEqual(ranking[0]["drag_type"], "偶发长尾")

    def test_build_daily_problem_list_prioritizes_recurring_drag(self):
        workflow_rows = [
            {
                "workflow_name": "E2E-Full",
                "run_count": 5,
                "affected_pr_count": 4,
                "total_runtime_minutes": 600.0,
                "avg_runtime_minutes": 120.0,
                "max_runtime_minutes": 180.0,
                "queue_p90_minutes": 10.0,
                "execution_p90_minutes": 140.0,
            }
        ]
        job_rows = [
            {
                "workflow_name": "E2E-Full",
                "job_name": "integration-tests",
                "run_count": 5,
                "max_runtime_minutes": 150.0,
                "avg_runtime_minutes": 120.0,
                "total_runtime_minutes": 600.0,
                "queue_p90_minutes": 10.0,
                "drag_type": "高频拖慢项",
            }
        ]
        step_rows = [
            {
                "workflow_name": "E2E-Full",
                "job_name": "integration-tests",
                "step_name": "run tests",
                "run_count": 5,
                "max_runtime_minutes": 100.0,
                "avg_runtime_minutes": 90.0,
                "total_runtime_minutes": 450.0,
            }
        ]

        rows = MODULE.build_daily_problem_list(workflow_rows, job_rows, step_rows, ["Anomaly"])

        self.assertEqual(rows[0]["Status"], "likely_recurring")
        self.assertIn(rows[0]["Type"], {"workflow", "job", "step"})

    def test_get_report_views_returns_precomputed_views(self):
        report = {
            "workflow_rows": [
                {"workflow_name": "E2E-Full", "run_count": 2, "total_runtime_minutes": 380.0},
            ],
            "job_rows": [
                {"workflow_name": "E2E-Full", "job_name": "integration-tests", "run_count": 2},
            ],
            "step_rows": [
                {"workflow_name": "E2E-Full", "job_name": "integration-tests", "step_name": "run tests", "run_count": 2},
            ],
        }

        workflow_rows, job_rows, step_rows = MODULE.get_report_views(report)

        self.assertEqual(workflow_rows[0]["workflow_name"], "E2E-Full")
        self.assertEqual(workflow_rows[0]["run_count"], 2)
        self.assertEqual(job_rows[0]["job_name"], "integration-tests")
        self.assertEqual(job_rows[0]["run_count"], 2)
        self.assertEqual(step_rows[0]["step_name"], "run tests")

    def test_write_excel_daily_mode_creates_daily_sheets(self):
        report = {
            "legacy_row": {
                "Repository": "owner/repo",
                "PR E2E时长 P90(min)": 0.0,
                "CI E2E时长 P90(min)": 0.0,
                "排队耗时 P90(min)": 0.0,
                "CI执行时长 P90(min)": 0.0,
                "PR检视时长 P90(min)": "Not reliable",
                "CI E2E达标率(%)": 0.0,
                "统计PR数": 0,
            },
            "repo": "owner/repo",
            "window": MODULE.Window("2026-04-19", "2026-04-19"),
            "pr_rows": [
                {
                    "number": 123,
                    "title": "Raw row demo",
                    "html_url": "https://example.com/pr/123",
                    "workflows": [
                        {
                            "run_id": 999,
                            "name": "CI",
                            "status": "completed",
                            "conclusion": "success",
                            "branch": "main",
                            "head_sha": "abc123",
                            "event": "pull_request",
                            "actor": "octocat",
                            "created_at": "2026-04-19T01:00:00Z",
                            "started_at": "2026-04-19T01:02:00Z",
                            "completed_at": "2026-04-19T01:20:00Z",
                            "queue_minutes": 2.0,
                            "execution_minutes": 18.0,
                            "html_url": "https://example.com/runs/999",
                            "jobs": [
                                {
                                    "job_id": 321,
                                    "name": "test",
                                    "check_run_url": "https://example.com/check-runs/1",
                                    "runner_id": 7,
                                    "runner_name": "runner-1",
                                    "runner_os": "Linux",
                                    "runner_arch": "X64",
                                    "runner_group": "default",
                                    "runner_labels": ["ubuntu-latest", "x64"],
                                    "status": "completed",
                                    "conclusion": "success",
                                    "created_at": "2026-04-19T01:00:00Z",
                                    "started_at": "2026-04-19T01:02:00Z",
                                    "completed_at": "2026-04-19T01:20:00Z",
                                    "queue_minutes": 2.0,
                                    "execution_minutes": 18.0,
                                    "html_url": "https://example.com/jobs/321",
                                    "steps": [
                                        {
                                            "number": 1,
                                            "name": "checkout",
                                            "status": "completed",
                                            "conclusion": "success",
                                            "started_at": "2026-04-19T01:02:00Z",
                                            "completed_at": "2026-04-19T01:03:00Z",
                                            "duration_minutes": 1.0,
                                            "raw_step_index": 1,
                                        }
                                    ],
                                    "steps_partial": False,
                                }
                            ],
                        }
                    ],
                }
            ],
            "daily_problems": [{
                "Type": "job",
                "Name": "wf / job",
                "Scope": "wf",
                "Run Count": 2,
                "Max Runtime(min)": 100.0,
                "Avg Runtime(min)": 80.0,
                "Total Runtime(min)": 160.0,
                "Status": "current_issue",
                "Why It Matters": "drag",
            }],
            "workflow_rows": [],
            "job_rows": [],
            "step_rows": [],
            "coverage_notes": [],
        }

        with NamedTemporaryFile(suffix=".xlsx") as temp:
            MODULE.write_excel([report], temp.name, "daily_diagnostic")
            workbook = load_workbook(temp.name, read_only=True)
            self.assertIn("CI效率报告", workbook.sheetnames)
            self.assertIn("Current Problems", workbook.sheetnames)
            self.assertIn("Daily Drill-down", workbook.sheetnames)
            self.assertIn("Workflow Raw", workbook.sheetnames)
            self.assertIn("Job Raw", workbook.sheetnames)
            self.assertIn("Step Raw", workbook.sheetnames)

            workflow_sheet = workbook["Workflow Raw"]
            workflow_headers = [cell.value for cell in next(workflow_sheet.iter_rows(min_row=1, max_row=1))]
            self.assertIn("workflow_run_id", workflow_headers)
            self.assertIn("workflow_name", workflow_headers)
            workflow_values = [cell.value for cell in next(workflow_sheet.iter_rows(min_row=2, max_row=2))]
            self.assertIn("CI", workflow_values)

            job_sheet = workbook["Job Raw"]
            job_values = [cell.value for cell in next(job_sheet.iter_rows(min_row=2, max_row=2))]
            self.assertIn("test", job_values)
            self.assertIn("ubuntu-latest, x64", job_values)
            self.assertIn("Linux", job_values)
            self.assertIn("https://example.com/check-runs/1", job_values)

            step_sheet = workbook["Step Raw"]
            step_values = [cell.value for cell in next(step_sheet.iter_rows(min_row=2, max_row=2))]
            self.assertIn("checkout", step_values)


if __name__ == "__main__":
    unittest.main()
