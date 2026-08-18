#!/usr/bin/env python3
"""
CI 效率分析报告生成器

从 Action Insight 本地 PostgreSQL 读取 GitHub Actions 数据，生成 CI 效率分析报告（Excel + HTML）。
支持按仓库配置进行多项目对比，全程不调用 GitHub API。

用法:
  # 单仓库（默认 vllm-ascend，最近 30 天）
  python ci_analyze.py

  # 指定仓库
  python ci_analyze.py --repo vllm-project/vllm-ascend

  # 多仓库对比
  python ci_analyze.py --repo vllm-project/vllm-ascend --repo vllm-project/vllm

  # 指定时间范围
  python ci_analyze.py --repo vllm-project/vllm-ascend --from 2026-05-01 --to 2026-05-23

  # 仅列出可用仓库
  python ci_analyze.py --list-repos

  # 自定义 step 分类映射
  python ci_analyze.py --step-names my-step-names.json

  # 只分析指定工作流（名称子串匹配，可多次指定）
  python ci_analyze.py --repo vllm-project/vllm-ascend --workflow "build" --workflow "test"

  # 跳过 Excel 输出（仅打印到终端）
  python ci_analyze.py --no-excel

  # 下钻 HTML（默认生成）：列出 >60min 的 run，点击下钻查看 job 条形图与 step 明细
  python ci_analyze.py
  python ci_analyze.py --drilldown-min 30        # 改阈值
  python ci_analyze.py --no-drilldown            # 跳过下钻 HTML
"""

import argparse
import json
import math
import sys
from collections import defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path

# ─── 配置 ──────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
SKILL_DIR = SCRIPT_DIR.parent
REPO_ROOT = SKILL_DIR.parents[2]
ENV_FILE = REPO_ROOT / ".env"
DEFAULT_CONFIG = REPO_ROOT / ".github-ci-efficiency.yaml"
DEFAULT_STEP_NAMES = SKILL_DIR / "references" / "step-names.json"
DEFAULT_PG_URL = "postgresql://action_insight:action_insight@localhost:5433/action_insight"
DEFAULT_DAYS = 30


def load_env(path: Path) -> dict:
    env = {}
    if path.exists():
        for line in path.read_text().splitlines():
            line = line.strip()
            if "=" in line and not line.startswith("#"):
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip('"')
    return env


# ─── 本地 PostgreSQL 客户端 ────────────────────────────────────────────

class PostgresClient:
    """只读 PostgreSQL 查询客户端。连接串来自 PG_DATABASE_URL。"""

    def __init__(self, db_url: str):
        try:
            import psycopg
            from psycopg.rows import dict_row
        except ImportError:
            print("[ERROR] 缺少 psycopg；请用 `uv run` 执行本脚本")
            sys.exit(1)
        self._psycopg = psycopg
        self._dict_row = dict_row
        self.db_url = db_url

    def query(self, sql: str) -> list[dict]:
        try:
            with self._psycopg.connect(self.db_url, row_factory=self._dict_row) as conn:
                conn.execute("SET TRANSACTION READ ONLY")
                return list(conn.execute(sql).fetchall())
        except self._psycopg.Error as e:
            print(f"[ERROR] PostgreSQL query failed: {e}")
            snippet = sql[:200] + "..." if len(sql) > 200 else sql
            print(f"  SQL: {snippet}")
            sys.exit(1)


# ─── 数据获取 ──────────────────────────────────────────────────────────

def get_repo_ids(client: PostgresClient) -> dict[str, int]:
    """Return {owner/repo: id} for all non-test repos."""
    rows = client.query(
        "SELECT id, owner, repo FROM repos WHERE owner NOT LIKE 'big-%' AND owner NOT LIKE 'perf-%' AND owner NOT LIKE 'test-%' ORDER BY id"
    )
    return {f"{r['owner']}/{r['repo']}": r["id"] for r in rows}


def _workflow_match_clause(patterns: list[str] | None) -> str:
    """runs.name 精确、不区分大小写匹配的 SQL AND 子句。"""
    if not patterns:
        return ""
    def esc(p: str) -> str:
        return p.replace("'", "''")
    clauses = " OR ".join(f"name ILIKE '{esc(p)}'" for p in patterns)
    return f" AND ({clauses})"


def _workflow_file_clause(files: list[str] | None) -> str:
    """workflow 文件名过滤的 SQL AND 子句。

    runs.workflow_file 在部分 DB 为空（ETL 未回填），workflow_attempts 表也未必覆盖
    所有 run。故同时匹配 workflow_file/workflow_path 列 + workflow_attempts 子查询。
    调用方还应配合 _resolve_workflow_names 动态拿显示名做 name 兜底过滤。
    """
    if not files:
        return ""
    def esc(p: str) -> str:
        return p.replace("'", "''")
    clauses = " OR ".join(
        f"workflow_file = '{esc(p)}' OR workflow_path LIKE '%/{esc(p)}'"
        f" OR id IN (SELECT run_id FROM workflow_attempts WHERE workflow_file = '{esc(p)}')"
        for p in files
    )
    return f" AND ({clauses})"


def _resolve_workflow_names(client, files: list[str]) -> list[str]:
    """从 DB 查 workflow_file→显示名映射，返回 name 列表供 name LIKE 兜底过滤。

    workflow_attempts JOIN runs 拿 file→name；查不到则返回空（调用方不过滤）。
    """
    if not files:
        return []
    def esc(p: str) -> str:
        return p.replace("'", "''")
    names = set()
    for f in files:
        rows = client.query(
            f"SELECT DISTINCT r.name FROM workflow_attempts wa JOIN runs r ON wa.run_id=r.id "
            f"WHERE wa.workflow_file = '{esc(f)}'"
        )
        for row in rows:
            n = row.get("name")
            if n and "/" not in n:  # 排除 path 脏值（如 .github/workflows/xxx）
                names.add(n)
    return list(names)


def parse_config_entries(path: str) -> dict[str, list[dict]]:
    """读取项目对比配置，保留 workflow 显示名和稳定文件名。"""
    try:
        import yaml
    except ImportError:
        print("[ERROR] 缺少 PyYAML；请用 `uv run` 执行本脚本")
        sys.exit(1)
    data = yaml.safe_load(Path(path).read_text(encoding="utf-8")) or {}
    return {
        item["repo"]: [w for w in item.get("workflows", []) if w.get("name")]
        for item in data.get("repositories", []) if item.get("repo")
    }


def parse_config(path: str) -> dict[str, list[str]]:
    """读取项目对比配置，返回 {owner/repo: [workflow display name]}。"""
    return {
        repo: [workflow["name"] for workflow in workflows]
        for repo, workflows in parse_config_entries(path).items()
    }


def normalize_configured_workflows(runs: list[dict], workflows: list[dict]) -> None:
    """用稳定 workflow_file 修正动态 run-name，供统计按 workflow 聚合。"""
    names_by_file = {w["file"]: w["name"] for w in workflows if w.get("file")}
    for run in runs:
        configured_name = names_by_file.get(run.get("workflow_file"))
        if configured_name:
            run["run_name"] = run.get("name")
            run["name"] = configured_name


def resolve_workflow_file(workflow_name: str, runs: list[dict], workflows: list[dict]) -> str:
    """配置未写 file 时，从命中 runs 中取最常见的 workflow_file。"""
    configured = next(
        (w.get("file") for w in workflows if w["name"] == workflow_name and w.get("file")),
        None,
    )
    if configured:
        return configured
    counts = defaultdict(int)
    for run in runs:
        if run.get("workflow_file"):
            counts[run["workflow_file"]] += 1
    return max(counts, key=counts.get) if counts else ""


def fetch_runs(client: PostgresClient, repo_ids: list[int], date_from: str, date_to: str, workflow_patterns: list[str] | None = None, workflow_files: list[str] | None = None) -> list[dict]:
    repo_id_list = ",".join(str(x) for x in repo_ids)
    clauses = [
        clause.removeprefix(" AND ")
        for clause in (
            _workflow_match_clause(workflow_patterns),
            _workflow_file_clause(workflow_files),
        ) if clause
    ]
    wf_clause = f" AND ({' OR '.join(clauses)})" if clauses else ""
    # Fix 2.2: 使用独占上界范围查询，允许查询优化器使用索引扫描
    date_to_next = (datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    return client.query(
        f"SELECT id, repo_id, name, head_branch, head_sha, event, "
        f"status, conclusion, created_at, updated_at, html_url, "
        f"duration_seconds, date, workflow_file "
        f"FROM runs "
        f"WHERE repo_id IN ({repo_id_list}) "
        f"AND date >= '{date_from}' AND date < '{date_to_next}'"
        f"{wf_clause} "
        f"ORDER BY created_at DESC"
    )


def fetch_jobs(client: PostgresClient, run_ids: list[int]) -> list[dict]:
    if not run_ids:
        return []
    all_jobs = []
    # 分批限制单次查询和返回集大小。
    # 缩小批次至 500 run_ids（≈ 2500 jobs），避免静默截断
    for i in range(0, len(run_ids), 500):
        batch = run_ids[i : i + 500]
        id_list = ",".join(str(x) for x in batch)
        jobs = client.query(
            f"SELECT id, run_id, name, status, conclusion, "
            f"created_at, started_at, completed_at, html_url, "
            f"queue_duration_seconds, duration_seconds, labels_json, "
            f"resource_model AS card_model, resource_count AS card_count "
            f"FROM jobs WHERE run_id IN ({id_list}) "
            f"UNION ALL "
            f"SELECT wj.job_id AS id, wj.run_id, wj.name, wj.status, wj.conclusion, "
            f"wj.created_at, wj.started_at, wj.completed_at, wj.html_url, "
            f"wj.queue_duration_seconds, wj.duration_seconds, wj.labels_json, "
            f"wj.resource_model AS card_model, wj.resource_count AS card_count "
            f"FROM workflow_jobs wj WHERE wj.run_id IN ({id_list}) "
            f"AND NOT EXISTS (SELECT 1 FROM jobs j WHERE j.id = wj.job_id)"
        )
        for job in jobs:
            try:
                job["labels"] = json.loads(job.get("labels_json") or "[]")
            except (TypeError, json.JSONDecodeError):
                job["labels"] = []
        all_jobs.extend(jobs)
    return all_jobs


def fetch_steps(client: PostgresClient, job_ids: list[int]) -> list[dict]:
    """Fetch steps for jobs in bounded batches."""
    if not job_ids:
        return []
    all_steps = []
    # 每批 500 个 job_id，避免响应过大被截断
    batch_size = 500
    total = len(job_ids)
    for i in range(0, total, batch_size):
        batch = job_ids[i : i + batch_size]
        id_list = ",".join(str(x) for x in batch)
        steps = client.query(
            f"SELECT job_id, number, name, status, conclusion, "
            f"started_at, completed_at, duration_seconds "
            f"FROM steps WHERE job_id IN ({id_list})"
        )
        all_steps.extend(steps)
        if (i // batch_size + 1) % 20 == 0 or i + batch_size >= total:
            print(f"    Steps 进度: {min(i + batch_size, total)}/{total} jobs ({len(all_steps)} steps fetched)")
    return all_steps


def fetch_pr_metrics(client: PostgresClient, repo_ids: list[int], date_from: str, date_to: str) -> list[dict]:
    repo_id_list = ",".join(str(x) for x in repo_ids)
    # Fix 2.2: 使用独占上界范围查询，允许查询优化器使用索引扫描
    date_to_next = (datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)).strftime("%Y-%m-%d")
    return client.query(
        f"SELECT id, repo_id, pr_number, title, branch, author, state, "
        f"html_url, created_at, ci_started_at, ci_completed_at, "
        f"merged_at, time_to_ci_start_seconds, ci_duration_seconds, "
        f"time_to_merge_seconds, merge_lead_time_seconds, "
        f"workflow_count, successful_workflow_count, conclusion "
        f"FROM pr_metrics "
        f"WHERE repo_id IN ({repo_id_list}) "
        f"AND created_at >= '{date_from}' AND created_at < '{date_to_next}' "
        f"ORDER BY created_at DESC"
    )


def fetch_pr_workflows(client: PostgresClient, pr_metric_ids: list[int], run_id_filter: set[int] | None = None) -> list[dict]:
    """Fetch PR-workflow links.

    run_id_filter: 若指定，则在 SQL 层过滤 run_id，避免 Python 端冗余过滤大量数据（Fix 2.3）。
    """
    if not pr_metric_ids:
        return []
    all_links = []
    # Fix 2.1: 缩小批次至 1000，减少单次结果集超限风险
    for i in range(0, len(pr_metric_ids), 1000):
        batch = pr_metric_ids[i : i + 1000]
        id_list = ",".join(str(x) for x in batch)
        # Fix 2.3: 将 run_id 过滤下推到 SQL 层
        run_filter = ""
        if run_id_filter:
            run_ids_sql = ",".join(str(x) for x in run_id_filter)
            run_filter = f" AND run_id IN ({run_ids_sql})"
        links = client.query(
            f"SELECT pr_metric_id, run_id "
            f"FROM pr_workflows WHERE pr_metric_id IN ({id_list}){run_filter}"
        )
        all_links.extend(links)
    return all_links


# ─── 统计工具 ──────────────────────────────────────────────────────────

def percentile(values: list[float], p: float) -> float | None:
    """Percentile, p in 0-1 range (0.5=P50, 0.9=P90). 统一 0-1 制（ADR-005）。"""
    clean = sorted(v for v in values if v is not None)
    if not clean:
        return None
    if len(clean) == 1:
        return round(clean[0], 3)
    k = (len(clean) - 1) * p
    f, c = math.floor(k), math.ceil(k)
    if f == c:
        return round(clean[f], 3)
    return round(clean[f] * (c - k) + clean[c] * (k - f), 3)


def sec_to_min(s) -> float | None:
    """Convert seconds to minutes, handling string values from DB."""
    if s is None:
        return None
    try:
        return round(float(s) / 60.0, 3)
    except (ValueError, TypeError):
        return None


def safe_div(a: float, b: float) -> float:
    return round(float(a) / float(b), 3) if b else 0.0


def _card_hours(job: dict) -> float | None:
    """NPU card-hours: actual execution hours times resolved accelerator count."""
    count = job.get("card_count")
    if not isinstance(count, int) or count <= 0:
        return None
    started, completed = job.get("started_at"), job.get("completed_at")
    if not started or not completed:
        return None
    try:
        start = datetime.fromisoformat(str(started).replace("Z", "+00:00"))
        end = datetime.fromisoformat(str(completed).replace("Z", "+00:00"))
        seconds = (end - start).total_seconds()
        return round(seconds / 3600 * count, 6) if seconds >= 0 else None
    except (ValueError, TypeError):
        return None


def _is_cpu_job(job: dict) -> bool:
    """A job whose runner labels indicate a CPU-only machine (no accelerator)."""
    labels = job.get("labels") or []
    return any(isinstance(l, str) and "cpu" in l.lower() for l in labels)


def _cpu_hours(job: dict) -> float | None:
    """CPU wall-clock hours for CPU-only jobs; NPU jobs return None."""
    if not _is_cpu_job(job):
        return None
    started, completed = job.get("started_at"), job.get("completed_at")
    if not started or not completed:
        return None
    try:
        start = datetime.fromisoformat(str(started).replace("Z", "+00:00"))
        end = datetime.fromisoformat(str(completed).replace("Z", "+00:00"))
        seconds = (end - start).total_seconds()
        return round(seconds / 3600, 6) if seconds >= 0 else None
    except (ValueError, TypeError):
        return None


def _model_summary(jobs: list[dict]) -> dict[str, int]:
    """Total card count per device model across jobs."""
    out: dict[str, int] = {}
    for j in jobs:
        model = j.get("card_model")
        count = j.get("card_count")
        if isinstance(model, str) and isinstance(count, int) and count > 0:
            out[model] = out.get(model, 0) + count
    return out


def _timing_causes(rjobs: list[dict]) -> list[dict]:
    """Forensic timing causes computed from already-collected run/job/step data.

    Ported from github-workflow-forensics; 0 additional API calls."""
    def _sec(a, b):
        if not a or not b: return None
        try:
            da = datetime.fromisoformat(str(a).replace("Z", "+00:00"))
            db = datetime.fromisoformat(str(b).replace("Z", "+00:00"))
            return (db - da).total_seconds() if db >= da else None
        except (ValueError, TypeError):
            return None

    findings = []
    # build job timing from the JSON payload jobs (with nested steps)
    jnorm = []
    for j in rjobs:
        jc = j.get("created")
        js = j.get("started")
        je = j.get("completed")
        queue = _sec(jc, js)
        execution = _sec(js, je)
        step_durs = [s.get("dur") for s in j.get("steps", []) if s.get("dur") is not None]
        covered = sum(s * 60 for s in step_durs)  # steps store dur in minutes
        overhead = max(0, (execution or 0) - covered) if execution else None
        jnorm.append({"name": j.get("name", ""), "queue": queue, "execution": execution,
                      "overhead": overhead, "steps": j.get("steps", [])})
    # wall clock from run
    ends = [j.get("completed") for j in rjobs if j.get("completed")]
    wall = sum((e2 - e1).total_seconds() for e1, e2 in [])  # placeholder
    # compute wall from min(created) to max(completed)
    starts = [j.get("created") for j in rjobs if j.get("created")]
    if starts and ends:
        try:
            t0 = min(datetime.fromisoformat(s.replace("Z", "+00:00")) for s in starts)
            t1 = max(datetime.fromisoformat(e.replace("Z", "+00:00")) for e in ends)
            wall = (t1 - t0).total_seconds()
        except (ValueError, TypeError):
            wall = None
    # longest queue
    lq = max(jnorm, key=lambda x: x["queue"] or -1, default=None)
    if lq and lq["queue"]:
        findings.append({"kind": "Runner queue", "subject": lq["name"], "dur": lq["queue"], "wall": wall, "evidence": "created_at -> started_at", "strength": "direct"})
    # longest execution
    le = max(jnorm, key=lambda x: x["execution"] or -1, default=None)
    if le and le["execution"]:
        findings.append({"kind": "Job execution", "subject": le["name"], "dur": le["execution"], "wall": wall, "evidence": "started_at -> completed_at", "strength": "direct"})
    # longest step
    all_steps = [(j, s) for j in jnorm for s in j["steps"] if s.get("dur") is not None]
    if all_steps:
        lj, ls = max(all_steps, key=lambda p: (p[1]["dur"] or 0) * 60)
        if ls["dur"]:
            findings.append({"kind": "Step execution", "subject": f"{lj['name']} / {ls['name']}", "dur": ls["dur"] * 60, "wall": wall, "evidence": "step timestamps", "strength": "direct"})
    # parallel tail
    completed_ends = sorted([e for e in ends if e], reverse=True)
    if len(completed_ends) > 1:
        try:
            t1 = datetime.fromisoformat(completed_ends[0].replace("Z", "+00:00"))
            t2 = datetime.fromisoformat(completed_ends[1].replace("Z", "+00:00"))
            tail_gap = (t1 - t2).total_seconds()
            if tail_gap >= 1:
                # find the job that finished last
                tail_job = next((j for j in rjobs if j.get("completed") == completed_ends[0]), {})
                findings.append({"kind": "Parallel tail", "subject": tail_job.get("name", ""), "dur": tail_gap, "wall": wall, "evidence": "gap between final and penultimate Job completion", "strength": "proxy"})
        except (ValueError, TypeError):
            pass
    # uninstrumented job time
    if jnorm:
        oh_job = max(jnorm, key=lambda x: x["overhead"] or -1)
        if oh_job["overhead"] and oh_job["overhead"] >= 1:
            findings.append({"kind": "Uninstrumented Job time", "subject": oh_job["name"], "dur": oh_job["overhead"], "wall": wall, "evidence": "Job execution minus summed Step durations", "strength": "proxy"})
    findings.sort(key=lambda f: -(f["dur"] or 0))
    return findings


def _calc_queue_min(job: dict, run: dict) -> float | None:
    """重算排队时间 = job.started_at - run.created_at（ADR: 修正 queue_duration_seconds 只算 job 内部等待的问题）。

    DB 预存的 queue_duration_seconds = job.started_at - job.created_at，只算了 runner 分配等待，
    漏掉了 run 创建→job 创建的等待（等上游 job）。真实排队应从 run 创建算起。
    """
    started = job.get("started_at")
    run_created = run.get("created_at")
    if not started or not run_created:
        return None
    try:
        s = datetime.fromisoformat(str(started).replace("Z", "+00:00"))
        r = datetime.fromisoformat(str(run_created).replace("Z", "+00:00"))
        diff = (s - r).total_seconds()
        return round(max(0, diff) / 60.0, 3)
    except (ValueError, TypeError):
        return None


# ─── 分析逻辑 ──────────────────────────────────────────────────────────

import re as _re

# step 分类正则兜底（ADR-005）：step-names.json 漏映射时用正则，避免"其他"失真
_STEP_RE_BUILD = _re.compile(r"install|checkout|cache|setup|compile|build|rebase|restore|get\s+(csrc|arch)|csrc", _re.I)
_STEP_RE_TEST = _re.compile(r"run\s+.*test|pre-commit|mypy|linkcheck|flake8|ruff|e2e", _re.I)
_STEP_RE_EXCLUDE = _re.compile(r"post\s|stop\s+(container|runner)|complete\s+job|clean\s+up", _re.I)


def classify_step(name: str, step_map: dict[str, str] | None = None) -> str:
    """优先查静态映射；未命中则正则兜底。统一两脚本的 step 分类（ADR-005）。"""
    if step_map:
        t = step_map.get(name)
        if t:
            return t
    if _STEP_RE_EXCLUDE.search(name):
        return "排除"
    if _STEP_RE_TEST.search(name):
        return "执行测试"
    if _STEP_RE_BUILD.search(name):
        return "构建"
    return "CI启动"


def _sp(prefix: str | None, name: str) -> str:
    """Sheet name with optional repo prefix, sanitized for Excel."""
    # Excel sheet names cannot contain: \ / * ? : [ ]
    safe_prefix = prefix.replace("/", "_").replace("\\", "").replace("*", "") if prefix else None
    if safe_prefix:
        combined = f"{safe_prefix} - {name}"
        return combined[:31]  # Excel sheet name limit
    return name[:31]


def _infer_resource_type(job_name: str) -> str:
    n = job_name.lower()
    if "310p" in n or "singlecard" in n or "single-card" in n or "single_node" in n:
        return "单卡 (310P)"
    if "multicard" in n or "multi-card" in n or "multi_node" in n:
        if "multicard-2" in n or "multi-card-2" in n or "2card" in n:
            return "2卡"
        if "4" in n:
            return "4卡"
        return "多卡"
    if "nvidia" in n or "cuda" in n or "gpu" in n:
        return "NVIDIA GPU"
    if "e2e" in n:
        return "E2E"
    if "unittest" in n or "unit_test" in n or "ut_" in n:
        return "单元测试"
    if "lint" in n or "flake8" in n or "ruff" in n or "mypy" in n:
        return "代码检查"
    return "其他"


def analyze_workflow_stats(runs, jobs, success_only=False, min_duration=0):
    # Fix 1.1: 拆分 run-level 和 job-level 聚合，避免运行次数/E2E 被 job 数量污染
    run_map = {r["id"]: r for r in runs}
    wf_groups = defaultdict(lambda: {
        "durations": [], "queues": [], "events": defaultdict(int), "created_ats": []
    })

    # 1. Run-level: duration、event、created_at
    for run in runs:
        # success_only: 只统计 success run 的耗时（ADR-005）
        if success_only and run.get("conclusion") != "success":
            continue
        # min_duration: 排除短 run（配合 success_only，目的2 的 >10min 过滤）
        if min_duration > 0:
            dur_chk = sec_to_min(run.get("duration_seconds"))
            if dur_chk is None or dur_chk <= min_duration:
                continue
        wf = run["name"]
        dur = sec_to_min(run.get("duration_seconds"))
        if dur is not None:
            wf_groups[wf]["durations"].append(dur)
        wf_groups[wf]["events"][run.get("event", "unknown")] += 1
        if run.get("created_at"):
            wf_groups[wf]["created_ats"].append(run["created_at"])

    # 2. Job-level: queue duration only
    for j in jobs:
        run = run_map.get(j["run_id"])
        if not run:
            continue
        q_dur = _calc_queue_min(j, run)
        if q_dur is not None:
            wf_groups[run["name"]]["queues"].append(q_dur)

    rows = []
    for wf, data in sorted(wf_groups.items(), key=lambda x: -len(x[1]["durations"])):
        durs = data["durations"]
        queues = data["queues"]
        dominant = max(data["events"], key=data["events"].get) if data["events"] else "unknown"

        # Fix 4.2: 调度周期 = 相邻 run 时间间隔的平均值，而非 duration 极差
        schedule_cycle = 0
        created_ats = data["created_ats"]
        if len(created_ats) >= 2:
            try:
                sorted_times = sorted(
                    datetime.fromisoformat(t.replace("Z", "+00:00"))
                    for t in created_ats if t
                )
                intervals = [
                    (sorted_times[i + 1] - sorted_times[i]).total_seconds() / 60
                    for i in range(len(sorted_times) - 1)
                ]
                schedule_cycle = round(sum(intervals) / len(intervals), 3) if intervals else 0
            except (ValueError, TypeError):
                schedule_cycle = 0

        rows.append({
            "工作流": wf,
            "触发类型": dominant,
            "运行次数": len(durs),  # 现在正确反映 run 数量
            "平均E2E(分钟)": safe_div(sum(durs), len(durs)),
            "P50 E2E(分钟)": percentile(durs, 0.5),
            "P90 E2E(分钟)": percentile(durs, 0.9),
            "平均排队(分钟)": safe_div(sum(queues), len(queues)),
            "P50 排队(分钟)": percentile(queues, 0.5),
            "P90 排队(分钟)": percentile(queues, 0.9),
            "调度周期(分钟)": schedule_cycle,
        })
    return rows


def analyze_job_stats(runs, jobs, success_only=False, min_duration=0):
    run_map = {r["id"]: r for r in runs}
    groups = defaultdict(lambda: {"durations": [], "queues": []})

    for j in jobs:
        # success_only: 只统计 success job 的耗时（ADR-005）
        if success_only and j.get("conclusion") != "success":
            continue
        run = run_map.get(j["run_id"])
        if not run:
            continue
        res_type = _infer_resource_type(j["name"])
        key = (f"{run['name']} / {j['name']}", res_type)
        d = sec_to_min(j.get("duration_seconds"))
        # min_duration: 排除耗时过短的样本，避免 0min job 污染统计
        if d is not None and d >= min_duration:
            groups[key]["durations"].append(d)
        q = _calc_queue_min(j, run)
        if q is not None:
            groups[key]["queues"].append(q)

    rows = []
    for (wf_job, res), data in sorted(groups.items(), key=lambda x: -len(x[1]["durations"])):
        durs = data["durations"]
        queues = data["queues"]
        rows.append({
            "工作流/任务": wf_job,
            "资源类型": res,
            "执行次数": len(durs),
            "平均E2E(分钟)": safe_div(sum(durs), len(durs)),
            "P50 E2E(分钟)": percentile(durs, 0.5),
            "P90 E2E(分钟)": percentile(durs, 0.9),
            "平均排队(分钟)": safe_div(sum(queues), len(queues)),
            "P50 排队(分钟)": percentile(queues, 0.5),
            "P90 排队(分钟)": percentile(queues, 0.9),
        })
    return rows


def analyze_step_stats(steps, step_names_map=None, success_only=False, min_duration=0):
    groups = defaultdict(lambda: {"durations": [], "success": 0, "total": 0})

    for s in steps:
        # success_only: 只统计 success step 的耗时（ADR-005）
        if success_only and s.get("conclusion") != "success":
            continue
        name = s["name"]
        stype = classify_step(name, step_names_map)
        if stype == "排除":
            continue
        key = (name, stype)
        d = sec_to_min(s.get("duration_seconds"))
        # min_duration: 排除耗时过短的样本（默认 0 仍排除 0/None）
        if d is not None and d >= min_duration:
            groups[key]["durations"].append(d)
        groups[key]["total"] += 1
        if s.get("conclusion") == "success":
            groups[key]["success"] += 1

    rows = []
    for (name, stype), data in sorted(groups.items(), key=lambda x: -sum(x[1]["durations"])):
        durs = data["durations"]
        rows.append({
            "步骤名称": name,
            "步骤类型": stype,
            "执行次数": data["total"],
            "平均耗时(分钟)": safe_div(sum(durs), len(durs)) if durs else 0,
            "P50 耗时(分钟)": percentile(durs, 0.5),
            "P90 耗时(分钟)": percentile(durs, 0.9),
            "成功率": round(data["success"] / data["total"] * 100, 1) if data["total"] else 0,
        })
    return rows


def _calc_review_min(pm):
    if pm.get("merged_at") and pm.get("ci_completed_at"):
        try:
            merged = datetime.fromisoformat(pm["merged_at"].replace("Z", "+00:00"))
            ci_done = datetime.fromisoformat(pm["ci_completed_at"].replace("Z", "+00:00"))
            return sec_to_min((merged - ci_done).total_seconds())
        except (ValueError, TypeError):
            pass
    return None


def _calc_pr_e2e_min(pm):
    """Fix 4.1: PR E2E = merged_at - created_at，覆盖完整 PR 生命周期（含评审等待）。

    原来使用 ci_duration_seconds 仅表示 CI 执行时长，语义不符。
    """
    if pm.get("merged_at") and pm.get("created_at"):
        try:
            merged = datetime.fromisoformat(pm["merged_at"].replace("Z", "+00:00"))
            created = datetime.fromisoformat(pm["created_at"].replace("Z", "+00:00"))
            return sec_to_min((merged - created).total_seconds())
        except (ValueError, TypeError):
            pass
    return None


def analyze_pr_stats(pr_metrics, pr_workflows):
    pr_wf_map = defaultdict(list)
    for pw in pr_workflows:
        pr_wf_map[pw["pr_metric_id"]].append(pw["run_id"])

    rows = []
    for pm in pr_metrics:
        wf_count = len(pr_wf_map.get(pm["id"], []))
        rows.append({
            "PR编号": pm.get("pr_number"),
            "标题": pm.get("title", ""),
            "作者": pm.get("author", ""),
            "创建时间": pm.get("created_at", ""),
            "合并时间": pm.get("merged_at") or "",
            "CI完成时间": pm.get("ci_completed_at") or "",
            # Fix 4.1: 使用 merged_at - created_at 计算完整 PR 生命周期
            "PR E2E(分钟)": _calc_pr_e2e_min(pm),
            "CI后评审(分钟)": _calc_review_min(pm),
            "工作流数量": wf_count,
            "链接": pm.get("html_url", ""),
            "CI结论": pm.get("conclusion", ""),
        })
    return rows


def build_pr_details(pr_metrics, pr_workflows, runs, jobs, steps):
    run_map = {r["id"]: r for r in runs}
    run_jobs = defaultdict(list)
    for j in jobs:
        run_jobs[j["run_id"]].append(j)
    job_steps = defaultdict(list)
    for s in steps:
        job_steps[s["job_id"]].append(s)
    pr_wf_map = defaultdict(list)
    for pw in pr_workflows:
        pr_wf_map[pw["pr_metric_id"]].append(pw["run_id"])

    _COMMON_KEYS = [
        "层级", "PR编号", "PR标题", "PR作者", "PR创建时间", "PR合并时间",
        "PR E2E(分钟)", "CI后评审(分钟)",
        "工作流名称", "工作流运行ID", "工作流状态", "工作流结论",
        "工作流创建时间", "工作流开始时间", "工作流完成时间", "工作流耗时(分钟)", "工作流排队(分钟)",
        "任务名称", "任务ID", "任务状态", "任务结论",
        "任务创建时间", "任务开始时间", "任务完成时间",
        "任务排队(分钟)", "任务耗时(分钟)",
        "步骤序号", "步骤名称", "步骤状态", "步骤结论",
        "步骤开始时间", "步骤完成时间", "步骤耗时(分钟)",
        "链接",
    ]

    def _base(level, pm, pr_e2e, review):
        return {k: None for k in _COMMON_KEYS} | {
            "层级": level,
            "PR编号": pm.get("pr_number"),
            "PR标题": pm.get("title", ""),
            "PR作者": pm.get("author", ""),
            "PR创建时间": pm.get("created_at", ""),
            "PR合并时间": pm.get("merged_at") or "",
            "PR E2E(分钟)": pr_e2e,
            "CI后评审(分钟)": review,
        }

    all_rows = []
    for pm in pr_metrics:
        # Fix 4.1: 使用 merged_at - created_at 计算完整 PR 生命周期
        pr_e2e = _calc_pr_e2e_min(pm)
        review = _calc_review_min(pm)
        wf_run_ids = pr_wf_map.get(pm["id"], [])

        all_rows.append(_base("PR", pm, pr_e2e, review) | {"链接": pm.get("html_url", "")})

        for run_id in wf_run_ids:
            run = run_map.get(run_id)
            if not run:
                continue
            wf_dur = sec_to_min(run.get("duration_seconds"))
            # 工作流排队 = 该 run 下最早 job 的 started_at - run.created_at
            rjobs = run_jobs.get(run_id, [])
            earliest_start = None
            for jj in rjobs:
                st = jj.get("started_at")
                if st and (earliest_start is None or st < earliest_start):
                    earliest_start = st
            wf_queue = _calc_queue_min({"started_at": earliest_start}, run) if earliest_start else None
            all_rows.append(_base("WORKFLOW", pm, pr_e2e, review) | {
                "工作流名称": run.get("name"),
                "工作流运行ID": run_id,
                "工作流状态": run.get("status"),
                "工作流结论": run.get("conclusion"),
                "工作流创建时间": run.get("created_at"),
                "工作流开始时间": run.get("created_at"),
                "工作流完成时间": run.get("updated_at"),
                "工作流耗时(分钟)": wf_dur,
                "工作流排队(分钟)": wf_queue,
                "链接": run.get("html_url", ""),
            })

            for j in run_jobs.get(run_id, []):
                all_rows.append(_base("JOB", pm, pr_e2e, review) | {
                    "工作流名称": run.get("name"),
                    "工作流运行ID": run_id,
                    "工作流状态": run.get("status"),
                    "工作流结论": run.get("conclusion"),
                    "工作流创建时间": run.get("created_at"),
                    "工作流开始时间": run.get("created_at"),
                    "工作流完成时间": run.get("updated_at"),
                    "工作流耗时(分钟)": wf_dur,
                    "任务名称": j.get("name"),
                    "任务ID": j.get("id"),
                    "任务状态": j.get("status"),
                    "任务结论": j.get("conclusion"),
                    "任务创建时间": j.get("created_at"),
                    "任务开始时间": j.get("started_at"),
                    "任务完成时间": j.get("completed_at"),
                    "任务排队(分钟)": _calc_queue_min(j, run),
                    "任务耗时(分钟)": sec_to_min(j.get("duration_seconds")),
                    "链接": j.get("html_url", ""),
                })

                # Fix 1.4: x.get("number", 0) 在 number 键存在但值为 None 时返回 None 而非 0
                for s in sorted(job_steps.get(j["id"], []), key=lambda x: x.get("number") if x.get("number") is not None else 0):
                    all_rows.append(_base("STEP", pm, pr_e2e, review) | {
                        "工作流名称": run.get("name"),
                        "工作流运行ID": run_id,
                        "工作流状态": run.get("status"),
                        "工作流结论": run.get("conclusion"),
                        "工作流创建时间": run.get("created_at"),
                        "工作流开始时间": run.get("created_at"),
                        "工作流完成时间": run.get("updated_at"),
                        "工作流耗时(分钟)": wf_dur,
                        "任务名称": j.get("name"),
                        "任务ID": j.get("id"),
                        "任务状态": j.get("status"),
                        "任务结论": j.get("conclusion"),
                        "任务创建时间": j.get("created_at"),
                        "任务开始时间": j.get("started_at"),
                        "任务完成时间": j.get("completed_at"),
                        "任务排队(分钟)": _calc_queue_min(j, run),
                        "任务耗时(分钟)": sec_to_min(j.get("duration_seconds")),
                        "步骤序号": s.get("number"),
                        "步骤名称": s.get("name"),
                        "步骤状态": s.get("status"),
                        "步骤结论": s.get("conclusion"),
                        "步骤开始时间": s.get("started_at"),
                        "步骤完成时间": s.get("completed_at"),
                        "步骤耗时(分钟)": sec_to_min(s.get("duration_seconds")),
                    })

    return all_rows


def analyze_comparison(repos_data: dict[str, dict]) -> list[dict]:
    rows = []
    for repo_name, data in repos_data.items():
        runs = data.get("runs", [])
        jobs = data.get("jobs", [])
        pr_metrics = data.get("pr_metrics", [])

        # Fix 1.4: 使用 sec_to_min() 代替裸 float()，避免空字符串或畸形数据崩溃
        wf_durs = [v for v in (sec_to_min(r.get("duration_seconds")) for r in runs) if v is not None]
        job_durs = [v for v in (sec_to_min(j.get("duration_seconds")) for j in jobs) if v is not None]
        # 重算排队：job.started_at - run.created_at
        run_map = {r["id"]: r for r in runs}
        job_queues = [v for v in (_calc_queue_min(j, run_map.get(j["run_id"], {})) for j in jobs) if v is not None]

        conclusions = defaultdict(int)
        for j in jobs:
            conclusions[j.get("conclusion") or "unknown"] += 1
        total = len(jobs)

        merged_prs = sum(1 for pm in pr_metrics if pm.get("merged_at"))
        ci_durations = [float(pm["ci_duration_seconds"]) / 60.0 for pm in pr_metrics if pm.get("ci_duration_seconds") is not None]

        events = defaultdict(int)
        for r in runs:
            events[r.get("event") or "unknown"] += 1

        rows.append({
            "仓库": repo_name,
            "总 Run 数": len(runs),
            "总 Job 数": total,
            "平均 Run 耗时(分钟)": safe_div(sum(wf_durs), len(wf_durs)) if wf_durs else 0,
            "P50 Run 耗时(分钟)": percentile(wf_durs, 0.5) if wf_durs else 0,
            "P90 Run 耗时(分钟)": percentile(wf_durs, 0.9) if wf_durs else 0,
            "平均 Job 耗时(分钟)": safe_div(sum(job_durs), len(job_durs)) if job_durs else 0,
            "P50 Job 耗时(分钟)": percentile(job_durs, 0.5) if job_durs else 0,
            "P90 Job 耗时(分钟)": percentile(job_durs, 0.9) if job_durs else 0,
            "平均排队(分钟)": safe_div(sum(job_queues), len(job_queues)) if job_queues else 0,
            "Job 成功率": round(conclusions.get("success", 0) / total * 100, 1) if total else 0,
            "Job 失败率": round(conclusions.get("failure", 0) / total * 100, 1) if total else 0,
            "PR 数量": len(pr_metrics),
            "已合并 PR": merged_prs,
            "平均 CI 时长(分钟)": safe_div(sum(ci_durations), len(ci_durations)) if ci_durations else 0,
            "主要触发类型": max(events, key=events.get) if events else "N/A",
        })
    return rows


# ─── 自动洞察 + HTML 输出（ADR-005: 从 ci_duration_analysis 合并）────────

def _norm_job_name(name: str) -> str:
    """去矩阵变体/分片号。复用 workflow_runs_on_date.norm_job_name 逻辑。"""
    import re
    s = re.sub(r'\s*card-\([^)]*\)', '', name)
    s = re.sub(r'\s*\([^)]*\)', '', s)
    s = re.sub(r'\s+', ' ', s).strip(' /')
    parts = [p.strip() for p in s.split('/') if p.strip()]
    if len(parts) >= 2:
        return f"{parts[0]} / {parts[-1]}"
    return parts[0] if parts else name


def generate_top_issues(runs, jobs, steps, step_map, min_duration=0):
    """返回 Top 问题列表，每项 {title, summary, evidence, severity}。

    接收 ci_analyze 的 dict 模型（runs/jobs/steps 平坦列表）。
    只分析 success job 的耗时，与 --success-only 口径一致。
    """
    issues = []
    n_runs = len(runs) or 1
    run_by_id = {r["id"]: r for r in runs}
    jobs_by_run = defaultdict(list)
    for j in jobs:
        jobs_by_run[j["run_id"]].append(j)

    longest_job_counter = defaultdict(int)
    job_dur = defaultdict(list)
    job_queue = defaultdict(list)
    step_dur = defaultdict(list)
    worst_queue_run = {}

    for rid, run_jobs in jobs_by_run.items():
        # min_duration: 短 job 不参与"最长 job"统计，避免 0min label job 被选为关键路径
        succ = [j for j in run_jobs if j.get("conclusion") == "success"
                and (sec_to_min(j.get("duration_seconds")) or 0) >= min_duration]
        if succ:
            longest = max(succ, key=lambda x: sec_to_min(x.get("duration_seconds")) or 0)
            longest_job_counter[_norm_job_name(longest["name"])] += 1
        for j in run_jobs:
            if j.get("conclusion") != "success":
                continue
            key = _norm_job_name(j["name"])
            dm = sec_to_min(j.get("duration_seconds"))
            if dm and dm >= min_duration:
                job_dur[key].append(dm)
            qm = _calc_queue_min(j, run_by_id.get(rid, {}))
            # 只对达到 min_duration 的 job 收集排队，避免短 job 的微小排队污染排队统计
            if qm is not None and dm and dm >= min_duration:
                job_queue[key].append(qm)
                if key not in worst_queue_run or qm > worst_queue_run[key][1]:
                    worst_queue_run[key] = (rid, qm)

    # steps 按成功 job 归属
    succ_job_ids = {j["id"] for j in jobs if j.get("conclusion") == "success"}
    for s in steps:
        if s.get("job_id") not in succ_job_ids:
            continue
        dm = sec_to_min(s.get("duration_seconds"))
        if dm and dm > 0:
            step_dur[s["name"]].append(dm)

    # 问题1: 排队瓶颈
    queue_issues = []
    for key, qs in job_queue.items():
        if len(qs) < 3:
            continue
        avg_q = sum(qs)/len(qs) if qs else 0
        p90_q = percentile(qs, 0.9) or 0
        avg_exec = sum(job_dur[key]) / len(job_dur[key]) if job_dur[key] else 0
        if (avg_exec > 0 and avg_q / avg_exec > 0.5) or p90_q > 30:
            queue_issues.append((key, avg_q, p90_q, avg_exec, len(qs)))
    if queue_issues:
        queue_issues.sort(key=lambda x: -x[2])
        k, aq, p90q, ae, cnt = queue_issues[0]
        ev_run, ev_q = worst_queue_run.get(k, (None, None))
        evidence = [
            f"{k}：平均排队 {aq:.0f}min / P90 排队 {p90q:.0f}min，但平均执行仅 {ae:.0f}min（{cnt} 次）",
            f"{k}：平均执行 {ae:.0f}min，排队占比 {aq/(ae+aq)*100:.0f}%" if (ae+aq) else "",
        ]
        if ev_run:
            evidence.append(f"典型证据：run {ev_run} 中该 job 排队 {ev_q:.0f}min 才开始")
        issues.append({
            "title": "排队瓶颈：硬件池容量不足",
            "summary": f"『{k}』平均排队 {aq:.0f}min、P90 排队 {p90q:.0f}min（执行才 {ae:.0f}min），是拉长 run 墙钟的头号根因。",
            "evidence": [e for e in evidence if e],
            "severity": 100,
        })

    # 问题2: 关键路径 job
    if longest_job_counter:
        top_cp, top_cp_cnt = max(longest_job_counter.items(), key=lambda x: x[1])
        rate = top_cp_cnt / n_runs * 100
        cp_durs = job_dur.get(top_cp, [])
        cp_avg = sum(cp_durs)/len(cp_durs) if cp_durs else 0
        cp_p90 = percentile(cp_durs, 0.9) or 0
        issues.append({
            "title": f"关键路径：{top_cp}",
            "summary": f"『{top_cp}』在 {top_cp_cnt}/{n_runs} 个 run（{rate:.0f}%）中是最长 job，矩阵并行下它决定 run 墙钟。",
            "evidence": [
                f"{top_cp}：平均执行 {cp_avg:.0f}min / P90 {cp_p90:.0f}min（{len(cp_durs)} 次）",
                f"作为最长 job 出现 {top_cp_cnt} 次，占比 {rate:.0f}%",
            ],
            "severity": 80,
        })

    # 问题3: 最耗时 step
    if step_dur:
        top_step, top_durs = max(step_dur.items(), key=lambda x: sum(x[1]))
        type_dur = defaultdict(list)
        for s in steps:
            if s.get("job_id") not in succ_job_ids:
                continue
            dm = sec_to_min(s.get("duration_seconds"))
            if dm and dm > 0:
                type_dur[classify_step(s["name"], step_map)].append(dm)
        type_totals = [(t, sum(v)) for t, v in type_dur.items() if t != "排除"]
        grand = sum(x[1] for x in type_totals) or 1
        test_pct = next((v / grand * 100 for t, v in type_totals if t == "执行测试"), 0)
        issues.append({
            "title": f"step 热点：{top_step}",
            "summary": f"『{top_step}』总耗时 {sum(top_durs):.0f}min、均 {sum(top_durs)/len(top_durs):.0f}min，执行测试类 step 占总耗时 {test_pct:.0f}%。",
            "evidence": [
                f"{top_step}：执行 {len(top_durs)} 次，均 {sum(top_durs)/len(top_durs):.0f}min，总 {sum(top_durs):.0f}min",
                "step 类型占比：执行测试 " + f"{test_pct:.0f}%、" + "、".join(f"{t} {v/grand*100:.0f}%" for t, v in type_totals if t != "执行测试"),
            ],
            "severity": 60,
        })

    issues.sort(key=lambda x: -x["severity"])
    return issues[:3]


def write_html_report(filepath, repo, date_from, date_to, runs, jobs, steps,
                      step_map, success_only, min_duration, api_info):
    """输出 HTML 洞察报告：Top 问题 + 汇总卡片 + 统计表（ADR-005）。"""
    import html as html_lib
    top_issues = generate_top_issues(runs, jobs, steps, step_map, min_duration=min_duration)
    job_rows = analyze_job_stats(runs, jobs, success_only=success_only, min_duration=min_duration)
    step_rows = analyze_step_stats(steps, step_map, success_only=success_only, min_duration=min_duration)

    # run 级汇总（success only）
    run_durs = [sec_to_min(r.get("duration_seconds")) for r in runs
                if r.get("conclusion") == "success" and sec_to_min(r.get("duration_seconds"))]
    if min_duration > 0:
        run_durs = [d for d in run_durs if d > min_duration]

    def _fmt(v):
        return f"{v:.1f}" if isinstance(v, float) else str(v) if v is not None else "-"

    cards = ""
    if run_durs:
        cards = f'''<div class="cards">
          <div class="card"><div class="num">{len(run_durs)}</div><div class="lab">命中 run 数</div></div>
          <div class="card"><div class="num">{_fmt(sum(run_durs)/len(run_durs))}</div><div class="lab">平均耗时(分钟)</div></div>
          <div class="card"><div class="num">{_fmt(percentile(run_durs, 0.5))}</div><div class="lab">P50(分钟)</div></div>
          <div class="card"><div class="num">{_fmt(percentile(run_durs, 0.9))}</div><div class="lab">P90(分钟)</div></div>
          <div class="card"><div class="num">{_fmt(max(run_durs))}</div><div class="lab">最大(分钟)</div></div>
        </div>'''

    issue_cards = []
    for idx, issue in enumerate(top_issues, 1):
        sev_cls = f"issue-sev{min(idx, 3)}"
        ev_lis = "".join(f"<li>{html_lib.escape(e)}</li>" for e in issue.get("evidence", []))
        issue_cards.append(f'''
        <div class="issue">
          <div class="issue-head {sev_cls}">#{idx} {html_lib.escape(issue['title'])}</div>
          <div class="issue-body">
            <div class="issue-summary">{html_lib.escape(issue['summary'])}</div>
            <ul class="issue-evidence">{ev_lis}</ul>
          </div>
        </div>''')
    issues_html = f'<h2>🚨 Top 问题（按严重度）</h2><div class="issues">{chr(10).join(issue_cards)}</div>' if issue_cards else ""

    def _table(title, rows, note=""):
        if not rows:
            return f"<h3>{html_lib.escape(title)}</h3><p class='muted'>无数据</p>"
        headers = list(rows[0].keys())
        th = "".join(f"<th>{html_lib.escape(str(h))}</th>" for h in headers)
        trs = []
        for r in rows[:50]:
            tds = "".join(f"<td>{html_lib.escape(_fmt(r.get(h)))}</td>" for h in headers)
            trs.append(f"<tr>{tds}</tr>")
        note_html = f"<p class='muted'>{html_lib.escape(note)}</p>" if note else ""
        return f"<h3>{html_lib.escape(title)}</h3>{note_html}<table><thead><tr>{th}</tr></thead><tbody>{''.join(trs)}</tbody></table>"

    doc = f'''<!DOCTYPE html>
<html lang="zh"><head><meta charset="utf-8"><title>CI 耗时分析 - {html_lib.escape(repo)}</title>
<style>
  body {{ font-family: -apple-system, "Segoe UI", Roboto, sans-serif; margin: 24px auto; max-width: 1200px; color: #1f2328; }}
  h1 {{ border-bottom: 2px solid #4472C4; padding-bottom: 8px; }}
  h2 {{ color: #4472C4; margin-top: 32px; }}
  h3 {{ margin-top: 24px; }}
  .meta {{ color: #6b7280; font-size: 14px; margin-bottom: 16px; }}
  .cards {{ display: flex; flex-wrap: wrap; gap: 12px; margin: 16px 0; }}
  .card {{ background: #f0f5ff; border-radius: 8px; padding: 12px 20px; min-width: 110px; text-align: center; }}
  .card .num {{ font-size: 24px; font-weight: 700; color: #2c5cc5; }}
  .card .lab {{ font-size: 12px; color: #6b7280; margin-top: 4px; }}
  .issues {{ display: flex; flex-direction: column; gap: 14px; margin: 16px 0 24px; }}
  .issue {{ border: 1px solid #e1e4e8; border-radius: 8px; overflow: hidden; }}
  .issue-head {{ padding: 12px 16px; color: #fff; font-weight: 600; font-size: 15px; }}
  .issue-sev1 {{ background: #dc2626; }} .issue-sev2 {{ background: #ea580c; }} .issue-sev3 {{ background: #ca8a04; }}
  .issue-body {{ padding: 12px 16px; background: #fff; }}
  .issue-summary {{ font-size: 14px; line-height: 1.7; margin-bottom: 8px; }}
  .issue-evidence {{ margin: 6px 0 0; padding-left: 20px; }}
  .issue-evidence li {{ font-size: 13px; color: #475569; line-height: 1.6; margin: 3px 0; }}
  table {{ border-collapse: collapse; width: 100%; margin: 8px 0 24px; font-size: 13px; }}
  th {{ background: #4472C4; color: #fff; padding: 8px 10px; text-align: left; }}
  td {{ border: 1px solid #e1e4e8; padding: 6px 10px; }}
  tbody tr:nth-child(even) {{ background: #f6f8fa; }}
  .muted {{ color: #6b7280; font-size: 13px; }}
  .filter {{ background: #e6f4ea; padding: 8px 14px; border-radius: 4px; display: inline-block; margin-bottom: 16px; }}
</style></head><body>
<h1>CI 耗时分析报告</h1>
<div class="meta">仓库：<b>{html_lib.escape(repo)}</b> ｜ 时间范围：{html_lib.escape(date_from)} ~ {html_lib.escape(date_to)}</div>
<div class="filter">过滤：success={'on' if success_only else 'off'} min-duration={min_duration}min</div>
{cards}
{issues_html}
<h2>📊 统计数据</h2>
{_table("Job 耗时统计（仅成功 job）", job_rows, "全量见 Excel")}
{_table("Step 耗时统计（仅成功 step）", step_rows, "全量见 Excel")}
<p class="muted">{api_info}</p>
</body></html>'''
    from pathlib import Path as _P
    _P(filepath).write_text(doc, encoding="utf-8")
    print(f"✅ HTML 洞察报告: {filepath}", file=sys.stderr)


# ─── 下钻 HTML（ADR-009: >min 分钟 run 列表 → job 耗时条形图 → step 明细）──

def build_drilldown_data(repos_data: dict, step_map: dict | None = None, min_minutes: float = 60) -> dict:
    """组装下钻 HTML 所需的紧凑数据结构。

    纯函数（除 sec_to_min/_calc_queue_min/classify_step 外无副作用），便于单测。
    提交人经 pr_workflows → pr_metrics.author 关联；非 PR run 无作者，返回空串。
    """
    run_author: dict[int, str] = {}
    for data in repos_data.values():
        pm_by_id = {pm["id"]: pm for pm in data.get("pr_metrics", [])}
        for pw in data.get("pr_workflows", []):
            pm = pm_by_id.get(pw["pr_metric_id"])
            if pm and pm.get("author"):
                run_author.setdefault(pw["run_id"], pm["author"])

    out = []
    all_runs: list[dict] = []
    stats: dict[str, dict] = {}
    # ponytail: 有效阈值 10min（<10min 视为无效脏样本，约定）；min_minutes 是表格显示阈值（默认 60）
    VALID_MIN = 10.0
    for repo, data in repos_data.items():
        jobs_by_run = defaultdict(list)
        for j in data.get("jobs", []):
            jobs_by_run[j["run_id"]].append(j)
        # all_runs: run-level summary for CSV export, no threshold filter
        for r in data.get("runs", []):
            rjobs = jobs_by_run.get(r["id"], [])
            all_runs.append({
                "repo": repo,
                "author": run_author.get(r["id"], ""),
                "created": r.get("created_at", ""),
                "updated": r.get("updated_at", ""),
                "wf": r.get("name", ""),
                "event": r.get("event", ""),
                "dur": sec_to_min(r.get("duration_seconds")),
                "card_hours": sum(v for v in (_card_hours(j) for j in rjobs) if v is not None),
                "cpu_hours": sum(v for v in (_cpu_hours(j) for j in rjobs) if v is not None),
                "card_models": _model_summary(rjobs),
                "status": r.get("status", ""),
                "conclusion": r.get("conclusion", ""),
                "url": r.get("html_url", ""),
            })
        steps_by_job = defaultdict(list)
        for s in data.get("steps", []):
            steps_by_job[s["job_id"]].append(s)
        for r in data.get("runs", []):
            dur = sec_to_min(r.get("duration_seconds"))
            if dur is None or dur <= min_minutes:
                continue
            # 按真实开始时间排序（chronological），揭示串/并行关系；缺失时间回退 created_at
            rjobs = sorted(
                jobs_by_run.get(r["id"], []),
                key=lambda j: (j.get("started_at") or j.get("created_at") or ""),
            )
            jobs_json = []
            for j in rjobs:
                jsteps = sorted(
                    steps_by_job.get(j["id"], []),
                    key=lambda s: s.get("number") if s.get("number") is not None else 9999,
                )
                jobs_json.append({
                    "name": j.get("name", ""),
                    "dur": sec_to_min(j.get("duration_seconds")),
                    "card_count": j.get("card_count"),
                    "card_hours": _card_hours(j),
                    "cpu_hours": _cpu_hours(j),
                    "queue": _calc_queue_min(j, r),
                    "created": j.get("created_at", ""),
                    "started": j.get("started_at", ""),
                    "completed": j.get("completed_at", ""),
                    "status": j.get("status", ""),
                    "conclusion": j.get("conclusion", ""),
                    "url": j.get("html_url", ""),
                    "steps": [{
                        "n": s.get("number"),
                        "name": s.get("name", ""),
                        "dur": sec_to_min(s.get("duration_seconds")),
                        "status": s.get("status", ""),
                        "conclusion": s.get("conclusion", ""),
                        "type": classify_step(s.get("name", ""), step_map),
                    } for s in jsteps],
                })
            out.append({
                "repo": repo,
                "author": run_author.get(r["id"], ""),
                "created": r.get("created_at", ""),
                "updated": r.get("updated_at", ""),
                "wf": r.get("name", ""),
                "event": r.get("event", ""),
                "dur": dur,
                "card_hours": sum(v for v in (_card_hours(j) for j in rjobs) if v is not None),
                "cpu_hours": sum(v for v in (_cpu_hours(j) for j in rjobs) if v is not None),
                "card_models": _model_summary(rjobs),
                "timing_causes": _timing_causes(jobs_json),
                "status": r.get("status", ""),
                "conclusion": r.get("conclusion", ""),
                "url": r.get("html_url", ""),
                "jobs": jobs_json,
            })
    out.sort(key=lambda x: -x["dur"])
    # 每仓统计：基于 >10min（有效）的样本；avg/p50/p90 与排队均基于有效样本。
    # 单 run 的排队 = 该 run 内各 job 排队的最大值（与 workflow 排队口径一致）。
    for repo, data in repos_data.items():
        _jobs_by_run = defaultdict(list)
        for j in data.get("jobs", []):
            _jobs_by_run[j["run_id"]].append(j)
        valid: list[float] = []
        queues: list[float] = []
        for r in data.get("runs", []):
            d = sec_to_min(r.get("duration_seconds"))
            if d is None or d <= VALID_MIN:
                continue
            valid.append(d)
            rq = [_calc_queue_min(j, r) for j in _jobs_by_run.get(r["id"], [])]
            rq = [q for q in rq if q is not None]
            if rq:
                queues.append(max(rq))
        card_hours_by_run = {
            r["id"]: sum(v for v in (_card_hours(j) for j in _jobs_by_run.get(r["id"], [])) if v is not None)
            for r in data.get("runs", [])
        }
        cpu_hours_by_run = {
            r["id"]: sum(v for v in (_cpu_hours(j) for j in _jobs_by_run.get(r["id"], [])) if v is not None)
            for r in data.get("runs", [])
        }
        npu_by_model: dict[str, float] = {}
        for j in data.get("jobs", []):
            model = j.get("card_model")
            ch = _card_hours(j)
            if isinstance(model, str) and ch is not None:
                npu_by_model[model] = npu_by_model.get(model, 0) + ch
        stats[repo] = {
            "npu_hours": sum(card_hours_by_run.values()),
            "npu_failure_hours": sum(
                card_hours_by_run[r["id"]] for r in data.get("runs", [])
                if r.get("conclusion") in ("failure", "cancelled")
            ),
            "cpu_hours": sum(cpu_hours_by_run.values()),
            "cpu_failure_hours": sum(
                cpu_hours_by_run[r["id"]] for r in data.get("runs", [])
                if r.get("conclusion") in ("failure", "cancelled")
            ),
            "p50": percentile(valid, 0.5),
            "p90": percentile(valid, 0.9),
            "q_p50": percentile(queues, 0.5),
            "q_p90": percentile(queues, 0.9),
            "pass_rate": safe_div(len(valid) - sum(1 for d in valid if d > min_minutes), len(valid)) if valid else 0,
            "npu_by_model": npu_by_model,
            "valid": len(valid),  # 有效运行数（>10min）
            "over60": sum(1 for d in valid if d > min_minutes),  # > 显示阈值（默认60min）
        }
    return {"from": None, "to": None, "min": min_minutes, "validMin": VALID_MIN, "stats": stats, "runs": out, "all_runs": all_runs}


def write_drilldown_html(filepath, repos_data, date_from, date_to, step_map, api_info, min_minutes=60):
    """输出单文件下钻 HTML：首页 >min 分钟 run 表格，下钻 job 条形图，再下钻 step 明细。

    原生 HTML/CSS + 极少 JS（表格展开 + 按需渲染），无外部依赖（ADR-009）。
    """
    import json as _json
    payload = build_drilldown_data(repos_data, step_map, min_minutes)
    payload["from"] = date_from
    payload["to"] = date_to
    n = len(payload["runs"])
    blob = _json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")

    doc = f'''<!DOCTYPE html>
<html lang="zh"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>CI 效率报告 - {date_from} ~ {date_to}</title>
<style>
  body {{ font-family: -apple-system, "Segoe UI", Roboto, sans-serif; margin: 20px auto; max-width: 1400px; color: #1f2328; }}
  h1 {{ border-bottom: 2px solid #4472C4; padding-bottom: 8px; }}
  h2 {{ color: #1f2328; margin: 0 0 10px; font-size: 18px; }}
  .meta {{ color: #6b7280; font-size: 14px; margin-bottom: 12px; }}
  .tabs {{ display: flex; flex-wrap: wrap; gap: 6px; border-bottom: 2px solid #4472C4; margin-bottom: 14px; }}
  .tab {{ cursor: pointer; border: 1px solid #c3cddb; border-bottom: none; background: #eef2f7; color: #475569; padding: 8px 16px; border-radius: 6px 6px 0 0; font-size: 14px; font-weight: 600; }}
  .tab.active {{ background: #4472C4; color: #fff; border-color: #4472C4; }}
  .repo-panel {{ margin-bottom: 24px; }}
  .stats {{ margin: 10px 0 14px; }}
  .stats-table {{ border-collapse: collapse; width: 100%; font-size: 13px; }}
  .stats-table th {{ background: #4472C4; color: #fff; padding: 6px 10px; text-align: center; white-space: nowrap; }}
  .stats-table td {{ border: 1px solid #e1e4e8; padding: 6px 10px; text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
  .stats-table .row-label {{ font-weight: 600; text-align: center; background: #f0f5ff; color: #2c5cc5; }}
  .stats-table .pass-cell {{ text-align: center; font-weight: 700; font-size: 16px; color: #2c5cc5; }}
  .model-breakdown {{ margin: 6px 0; font-size: 12px; color: #475569; }}
  .timing-causes {{ margin: 10px 0; padding: 8px 0; }}
  .timing-causes h4 {{ font-size: 12px; color: #6b7280; margin: 0 0 6px; text-transform: uppercase; letter-spacing: 0.08em; }}
  .tc-list {{ display: flex; flex-wrap: wrap; gap: 8px; }}
  .tc-item {{ display: grid; grid-template-columns: auto 1fr auto auto; gap: 6px; align-items: center; background: #f6f8fa; border: 1px solid #e1e4e8; border-radius: 4px; padding: 4px 10px; font-size: 12px; }}
  .tc-kind {{ font-weight: 600; color: #2c5cc5; }}
  .tc-subject {{ color: #374151; }}
  .tc-dur {{ font: 700 13px ui-monospace, monospace; color: #dc2626; }}
  .tc-pct {{ color: #6b7280; font-size: 11px; }}
  .tc-ev {{ grid-column: 1 / -1; color: #9ca3af; font-size: 10px; }}
  .table-wrap {{ overflow-x: auto; }}
  .table-wrap > table {{ min-width: 1300px; }}
  table {{ border-collapse: collapse; width: 100%; font-size: 13px; margin-bottom: 8px; }}
  th {{ background: #4472C4; color: #fff; padding: 8px 10px; text-align: left; white-space: nowrap; }}
  td {{ border: 1px solid #e1e4e8; padding: 6px 10px; vertical-align: top; }}
  tbody tr:nth-child(even) {{ background: #f6f8fa; }}
  .toggle {{ cursor: pointer; user-select: none; width: 28px; text-align: center; color: #4472C4; font-weight: 700; }}
  .arrow {{ display: inline-block; transition: transform .12s; }}
  .arrow.open {{ transform: rotate(90deg); }}
  tr.detail {{ display: none; }}
  tr.detail > td {{ background: #fff; padding: 14px 16px; }}
  .num {{ text-align: right; font-variant-numeric: tabular-nums; white-space: nowrap; }}
  .pill {{ border-radius: 99px; padding: 2px 9px; font-size: 11px; font-weight: 600; color: #fff; }}
  .pill.success {{ background: #16a34a; }} .pill.failure {{ background: #dc2626; }}
  .pill.cancelled {{ background: #6b7280; }} .pill.in_progress {{ background: #2563eb; }}
  a {{ color: #2c5cc5; }}
  /* job Gantt 甘特图（统一时间轴 + 时刻刻度） */
  .gantt {{ background: #fff; min-width: 1260px; padding: 4px 0; }}
  .gantt-meta {{ color: #6b7280; font-size: 12px; margin: 2px 0 6px; }}
  .gantt-ruler, .gjob > summary {{ display: grid; grid-template-columns: 180px minmax(120px, 1fr) 220px; align-items: center; gap: 8px; padding: 4px 8px; }}
  .gantt-ruler {{ border-bottom: 1px solid #d1d5db; color: #6b7280; font: 11px ui-monospace, monospace; }}
  .gantt-track {{ position: relative; height: 20px; border-radius: 3px; min-width: 120px; }}
  .gtick {{ position: absolute; top: 0; transform: translateX(-50%); font-size: 10px; color: #6b7280; white-space: nowrap; }}
  .gtick::after {{ content: ''; position: absolute; top: 14px; left: 50%; width: 1px; height: 4px; background: #9ca3af; }}
  .gjob {{ border-bottom: 1px solid #eef2f7; }}
  .gjob:last-of-type {{ border-bottom: none; }}
  .gjob > summary {{ cursor: pointer; list-style: none; }}
  .gjob > summary::-webkit-details-marker {{ display: none; }}
  .gjob-label {{ overflow: hidden; text-overflow: ellipsis; white-space: nowrap; font-size: 12px; text-decoration: none; color: #2c5cc5; }}
  .gjob-label:hover {{ text-decoration: underline; }}
  .gjob-dur {{ text-align: right; font-variant-numeric: tabular-nums; color: #374151; font-size: 11px; white-space: nowrap; }}
  .gjob[open] > summary {{ background: #f0f5ff; }}
  .bar {{ position: absolute; top: 3px; height: 14px; border-radius: 2px; min-width: 2px; box-shadow: 0 0 0 1px rgba(0,0,0,.08) inset; }}
  .bar.queue {{ background: #f0a33a; }}
  .bar.run {{ background: #4472C4; }}
  .bar[data-tip]:hover::after {{ content: attr(data-tip); position: absolute; left: 50%; bottom: 120%; transform: translateX(-50%); white-space: nowrap; background: #1f2328; color: #fff; font: 11px ui-monospace, monospace; padding: 4px 8px; border-radius: 4px; pointer-events: none; z-index: 10; }}
  .missing {{ position: absolute; left: 6px; top: 4px; color: #d84a3a; font: 700 10px ui-monospace, monospace; }}
  .steps {{ overflow-x: auto; padding: 6px 10px 10px; }}
  .steps table {{ font-size: 12px; table-layout: fixed; }}
  .steps td:nth-child(2) {{ overflow-wrap: anywhere; }}
  .steps th {{ background: #6b7280; }}
  .muted {{ color: #6b7280; font-size: 13px; }}
  .legend {{ font-size: 12px; color: #6b7280; margin: 6px 0 10px; display: flex; gap: 18px; }}
  .legend i {{ display: inline-block; width: 14px; height: 10px; margin-right: 5px; vertical-align: middle; border-radius: 2px; }}
  .btn-export {{ cursor: pointer; background: #4472C4; color: #fff; border: none; border-radius: 4px; padding: 6px 16px; font-size: 13px; font-weight: 600; }}
  .btn-export:hover {{ background: #2c5cc5; }}
</style></head><body>
<h1>CI 效率报告</h1>
<div class="meta">时间范围：<b>{date_from} ~ {date_to}</b> ｜ 阈值：&gt;{min_minutes}min ｜ 命中 <b>{n}</b> 个 run</div>
<div class="legend"><span><i style="background:#f0a33a"></i>排队</span><span><i style="background:#4472C4"></i>运行</span><span style="color:#9ca3af">点 job 展开 step</span></div>
<div class="tabs" id="tabs"></div>
<div id="panels"></div>
<p class="muted">{api_info}</p>
<script>const DATA={blob};
function esc(s){{return String(s==null?'':s).replace(/&/g,'&amp;').replace(/</g,'&lt;');}}
function fmt(v){{return v==null?'-':(typeof v==='number'?v.toFixed(1):v);}}
function fmtDurMS(ms){{const m=(ms||0)/60000;return isNaN(m)||m<0?'-':m.toFixed(1)+'min';}}
function fmtT(iso){{const d=new Date(iso);return isNaN(d)?'-':d.getHours().toString().padStart(2,'0')+':'+d.getMinutes().toString().padStart(2,'0')+':'+d.getSeconds().toString().padStart(2,'0');}}
function pill(c){{return '<span class="pill '+(c||'')+'">'+esc(c||'-')+'</span>';}}
const REPOS=Object.keys(DATA.stats||{{}});
const BY_REPO=REPOS.map(repo=>DATA.runs.filter(r=>r.repo===repo));
let activeRepo=0;
function renderTabs(){{let h='';REPOS.forEach((repo,i)=>{{h+='<button class="tab'+(i===activeRepo?' active':'')+'" onclick="selectTab('+i+')">'+esc(repo)+'</button>';}});document.getElementById('tabs').innerHTML=h;}}
function selectTab(i){{activeRepo=i;renderTabs();renderPanels();}}
function renderPanels(){{
  let h='';
  BY_REPO.forEach((runs,ri)=>{{
    h+='<div class="repo-panel" id="panel'+ri+'" style="display:'+(ri===activeRepo?'block':'none')+'">';
    h+='<h2>'+esc(REPOS[ri])+' CI效率报告</h2>';
    h+=renderStats(REPOS[ri]);
    h+='<div class="table-wrap"><table><thead><tr><th class="toggle"></th><th>代码仓</th><th>提交人</th><th>创建时间</th><th>结束时间</th><th>Workflow</th><th>耗时(min)</th><th>NPU卡时</th><th>CPU耗时</th><th>状态</th><th>Run URL</th></tr></thead><tbody id="rows'+ri+'"></tbody></table>'
    +'<div style="margin:8px 0"><button class="btn-export" onclick="exportCSV('+ri+')">导出 CSV</button></div></div>';
  }});
  document.getElementById('panels').innerHTML=h;
  BY_REPO.forEach((runs,ri)=>renderRows(ri));
}}
function renderStats(repo){{
  const s=DATA.stats&&DATA.stats[repo];
  if(!s)return '';
  const pct=(v)=>(v==null?'-':(v*100).toFixed(0)+'%');
  return '<div class="stats"><table class="stats-table">'
    +'<thead><tr><th></th><th>总机时</th><th>失败机时</th><th>P50耗时</th><th>P90耗时</th><th>P50排队</th><th>P90排队</th><th>达标率</th></tr></thead>'
    +'<tbody>'
    +'<tr><td class="row-label">NPU</td><td>'+fmt(s.npu_hours)+'</td><td>'+fmt(s.npu_failure_hours)+'</td>'
    +'<td class="pass-cell" rowspan="2">'+fmt(s.p50)+'</td><td class="pass-cell" rowspan="2">'+fmt(s.p90)+'</td><td class="pass-cell" rowspan="2">'+fmt(s.q_p50)+'</td><td class="pass-cell" rowspan="2">'+fmt(s.q_p90)+'</td><td class="pass-cell" rowspan="2">'+pct(s.pass_rate)+'</td></tr>'
    +'<tr><td class="row-label">CPU</td><td>'+fmt(s.cpu_hours)+'</td><td>'+fmt(s.cpu_failure_hours)+'</td></tr>'
    +'</tbody></table>'
    +(s.npu_by_model&&Object.keys(s.npu_by_model).length?'<div class="model-breakdown">型号分布: '+Object.entries(s.npu_by_model).sort((a,b)=>b[1]-a[1]).map(([m,h])=>m+' '+fmt(h)+'卡时').join(' ｜ ')+'</div>':'')
    +'</div>';
}}
function renderRows(ri){{
  const runs=BY_REPO[ri];let h='';
  runs.forEach((r,li)=>{{
    h+='<tr class="run-row"><td class="toggle" onclick="toggleRun('+ri+','+li+')"><span class="arrow" id="ar'+ri+'_'+li+'">▶</span></td>'
      +'<td>'+esc(r.repo)+'</td><td>'+(r.author?esc(r.author):'<span class="muted">'+esc(r.event||'-')+'</span>')+'</td>'
      +'<td>'+esc(r.created)+'</td><td>'+esc(r.updated)+'</td><td>'+esc(r.wf)+'</td><td class="num">'+r.dur.toFixed(1)+'</td><td class="num">'+fmt(r.card_hours)+'</td><td class="num">'+fmt(r.cpu_hours)+'</td>'
      +'<td>'+pill(r.conclusion||r.status)+'</td><td><a href="'+esc(r.url)+'" target="_blank">打开 ↗</a></td></tr>'
      +'<tr class="detail" id="det'+ri+'_'+li+'"><td colspan="11" id="dc'+ri+'_'+li+'"></td></tr>';
  }});
  document.getElementById('rows'+ri).innerHTML=h;
}}
function toggleRun(ri,li){{
  const det=document.getElementById('det'+ri+'_'+li),ar=document.getElementById('ar'+ri+'_'+li);
  const open=det.style.display==='table-row';
  if(open){{det.style.display='none';ar.classList.remove('open');return;}}
  if(!det.dataset.d){{document.getElementById('dc'+ri+'_'+li).innerHTML=renderJobs(ri,li);det.dataset.d='1';}}
  det.style.display='table-row';ar.classList.add('open');
}}
function renderJobs(ri,li){{
  const r=BY_REPO[ri][li];if(!r.jobs.length)return '<p class="muted">无 job 数据</p>';
  // 共享时间轴：起=run 触发(created)，止=所有 job 完成(max completed)
  const ends=r.jobs.map(j=>j.completed).filter(Boolean).sort();
  let aStart=r.created||r.jobs.map(j=>j.created).filter(Boolean).sort()[0];
  let aEnd=ends[ends.length-1]||r.updated||aStart;
  const t0=Date.parse(aStart),t1=Date.parse(aEnd),span=(t1-t0)||1;
  // 自适应刻度间隔：≈8-12 根线
  const hrs=span/3600000;
  let step=3600000;
  if(hrs>12)step=2*3600000; if(hrs>36)step=4*3600000; if(hrs>96)step=8*3600000; if(hrs>240)step=24*3600000;
  // 时刻刻度 + 网格线（对齐到所有 track）
  let ticks=[],stops=[];
  for(let t=Math.ceil(t0/step)*step;t<=t1;t+=step){{
    const p=(t-t0)/span*100; if(p<0.5||p>99.5)continue;
    ticks.push({{p:p,l:new Date(t).toLocaleTimeString('zh-CN',{{hour:'2-digit',minute:'2-digit'}})}});
    stops.push(p);
  }}
  let gp=['transparent 0'];
  stops.forEach(p=>{{p=Math.max(0,Math.min(100,p));gp.push('#e5e7eb '+(p-0.12).toFixed(2)+'%','#e5e7eb '+(p+0.12).toFixed(2)+'%','transparent '+(p+0.12).toFixed(2)+'%');}});
  gp.push('transparent 100%');
  const gb='linear-gradient(90deg,'+gp.join(',')+')';
  // 标尺行
  let ruler='<div class="gantt-ruler"><span></span><div class="gantt-track" style="background:'+gb+'">';
  ticks.forEach(t=>{{ruler+='<span class="gtick" style="left:'+t.p.toFixed(2)+'%">'+esc(t.l)+'</span>';}});
  ruler+='</div><span></span></div>';
  // 作业按真实开始时间排序（chronological），揭示串/并行
  const jobs=[...r.jobs].sort((a,b)=>(Date.parse(a.started||a.created||0))-(Date.parse(b.started||b.created||0)));
  let rows='';
  jobs.forEach((j)=>{{
    const jc=Date.parse(j.created),js=Date.parse(j.started),je=Date.parse(j.completed);
    const qL=(jc&&jc>=t0)?((jc-t0)/span*100):null;
    const qW=(js&&jc&&jc>=t0)?((js-jc)/span*100):0;
    const rL=(js&&js>=t0)?((js-t0)/span*100):null;
    const rW=(js&&je&&je>=js)?((je-js)/span*100):0;
    let bars='';
    const tip='启动 '+fmtT(j.started)+'  结束 '+fmtT(j.completed)+'  排队 '+fmtDurMS(js-jc)+'  运行 '+fmtDurMS(je-js);
    if(qL!=null)bars+='<span class="bar queue" style="left:'+qL.toFixed(2)+'%;width:'+Math.max(0.3,qW).toFixed(2)+'%" data-tip="'+esc(tip)+'"></span>';
    if(rL!=null)bars+='<span class="bar run" style="left:'+rL.toFixed(2)+'%;width:'+Math.max(0.3,rW).toFixed(2)+'%" data-tip="'+esc(tip)+'"></span>';
    if(!bars)bars='<span class="missing">时间缺失</span>';
    const jlabel=j.url?'<a class="gjob-label" href="'+esc(j.url)+'" target="_blank" rel="noopener" title="'+esc(j.name)+' (打开 job)" onclick="event.stopPropagation()">'+esc(j.name)+'</a>':'<span class="gjob-label" title="'+esc(j.name)+'">'+esc(j.name)+'</span>';
    rows+='<details class="gjob"><summary>'+jlabel
      +'<div class="gantt-track" style="background:'+gb+'">'+bars+'</div>'
      +'<span class="gjob-dur">排队 '+fmtDurMS(js-jc)+' · 运行 '+fmtDurMS(je-js)+' · '+(j.card_count!=null?('NPU卡时 '+fmt(j.card_hours)):('CPU耗时 '+fmt(j.cpu_hours)))+'</span></summary>'
      +renderSteps(j)+'</details>';
  }});
  let tc='';
  if(r.timing_causes&&r.timing_causes.length){{
    tc='<div class="timing-causes"><h4>Timing Causes</h4><div class="tc-list">';
    r.timing_causes.forEach(c=>{{
      const pct=c.wall?((c.dur/c.wall*100).toFixed(0)+'%'):'-';
      tc+='<div class="tc-item"><span class="tc-kind">'+esc(c.kind)+'</span><span class="tc-subject">'+esc(c.subject)+'</span><span class="tc-dur">'+fmtDurMS(c.dur*1000)+'</span><span class="tc-pct">'+pct+'</span><span class="tc-ev">'+esc(c.evidence)+'</span></div>';
    }});
    tc+='</div></div>';
  }}
  return '<div class="gantt"><div class="gantt-meta">时间轴：'+fmtT(aStart)+' → '+fmtT(aEnd)+'（共 '+((t1-t0)/60000).toFixed(0)+' min）</div>'+ruler+'<div class="gantt-body">'+rows+'</div>'+tc+'</div>';
}}
function renderSteps(j){{
  if(!j.steps.length)return '<p class="muted" style="padding:6px 10px">无 step 数据</p>';
  let h='<div class="steps"><table><thead><tr><th>#</th><th>步骤名称</th><th>类型</th><th>耗时(min)</th><th>状态</th></tr></thead><tbody>';
  j.steps.forEach(s=>{{h+='<tr><td>'+esc(s.n)+'</td><td>'+esc(s.name)+'</td><td>'+esc(s.type)+'</td>'
      +'<td class="num">'+fmt(s.dur)+'</td><td>'+pill(s.conclusion||s.status)+'</td></tr>';}});
  h+='</tbody></table></div>';return h;
}}
function csvCell(v){{v=String(v==null?'':v);return v.includes(',')||v.includes('"')||v.includes('\\n')?'"'+v.replace(/"/g,'""')+'"':v;}}
function exportCSV(ri){{
  const repo=REPOS[ri];const s=DATA.stats&&DATA.stats[repo]||{{}};
  const runs=DATA.all_runs||[];
  const rows=runs.filter(r=>r.repo===repo);
  let csv='';
  // stats header
  csv+='# 统计\\n';
  csv+='NPU总机时,'+csvCell(s.npu_hours)+'\\n';
  csv+='NPU失败机时,'+csvCell(s.npu_failure_hours)+'\\n';
  csv+='CPU总机时,'+csvCell(s.cpu_hours)+'\\n';
  csv+='CPU失败机时,'+csvCell(s.cpu_failure_hours)+'\\n';
  csv+='P50耗时,'+csvCell(s.p50)+'\\n';
  csv+='P90耗时,'+csvCell(s.p90)+'\\n';
  csv+='P50排队,'+csvCell(s.q_p50)+'\\n';
  csv+='P90排队,'+csvCell(s.q_p90)+'\\n';
  csv+='达标率,'+csvCell(s.pass_rate)+'\\n';
  csv+='# run 明细 ('+rows.length+' 条)\\n';
  csv+='代码仓,提交人,创建时间,结束时间,Workflow,触发事件,耗时(min),NPU卡时,CPU耗时,型号卡数,状态,结论,Run URL\\n';
  rows.forEach(r=>{{
    const ms=r.card_models?Object.entries(r.card_models).map(([m,c])=>m+'x'+c).join(' '):'';
    csv+=[r.repo,r.author,r.created,r.updated,r.wf,r.event,r.dur,r.card_hours,r.cpu_hours,ms,r.status,r.conclusion,r.url].map(csvCell).join(',')+'\\n';
  }});
  const blob=new Blob(['\uFEFF'+csv],{{type:'text/csv;charset=utf-8'}});
  const a=document.createElement('a');
  a.href=URL.createObjectURL(blob);
  a.download=repo.replace('/','_')+'-'+DATA.from+'_to_'+DATA.to+'.csv';
  a.click();
  URL.revokeObjectURL(a.href);
}}
renderTabs();renderPanels();
</script>
</body></html>'''
    from pathlib import Path as _P
    _P(filepath).write_text(doc, encoding="utf-8")
    print(f"✅ CI 效率报告: {filepath} ({n} runs)", file=sys.stderr)


# ─── Excel 输出 ─────────────────────────────────────────────────────────

def collect_workflow_job_queues(jobs: list[dict], run_map: dict, run_ids: set[int]) -> list[float]:
    """Queue samples from successful jobs, regardless of final workflow conclusion."""
    return [
        queue for queue in (
            _calc_queue_min(job, run_map.get(job["run_id"], {}))
            for job in jobs
            if job["run_id"] in run_ids and job.get("conclusion") == "success"
        ) if queue is not None
    ]


def overview_missing_reason(data: dict, durations: list, queues: list) -> str:
    """Explain every intentionally blank overview metric with source counts."""
    total_runs = data.get("total_run_count", 0)
    success_runs = data.get("success_run_count", 0)
    total_jobs = data.get("total_job_count", 0)
    success_jobs = data.get("success_job_count", 0)
    reasons = []
    if total_runs == 0:
        return "窗口内无 Run，E2E 与排队均无法计算"
    if not durations:
        if success_runs == 0:
            reasons.append(f"总 Run={total_runs}，成功 Run=0，E2E 不计算")
        else:
            threshold = data.get("min_duration", 0)
            reasons.append(f"成功 Run={success_runs}，但无耗时≥{threshold:g}分钟的有效样本")
    if not queues:
        if total_jobs == 0:
            reasons.append("Jobs=0，缺少排队时间源数据")
        elif success_jobs == 0:
            reasons.append(f"Jobs={total_jobs}，成功 Jobs=0，排队不计算")
        else:
            reasons.append(f"成功 Jobs={success_jobs}，但缺少可用 started_at/created_at")
    return "；".join(reasons)


def build_overview(overview_data: list[dict]) -> list[dict]:
    """总览页：每个仓库/workflow 一行，分别展示 E2E 和排队分布。"""
    rows = []
    for d in overview_data:
        durs = d.get("durations", [])
        queues = d.get("queues", [])
        rows.append({
            "仓库": d["repo"],
            "Workflow": d.get("workflow_file", ""),
            "Workflow显示名": d.get("workflow_name", ""),
            "总Run数": d.get("total_run_count", len(durs)),
            "成功Run数": d.get("success_run_count", len(durs)),
            "有效成功Run数": len(durs),
            "E2E P50(分钟)": percentile(durs, 0.5),
            "E2E 平均(分钟)": safe_div(sum(durs), len(durs)) if durs else None,
            "E2E P90(分钟)": percentile(durs, 0.9),
            "排队 P50(分钟)": percentile(queues, 0.5),
            "排队 平均(分钟)": safe_div(sum(queues), len(queues)) if queues else None,
            "排队 P90(分钟)": percentile(queues, 0.9),
            "空值判断依据": overview_missing_reason(d, durs, queues),
        })
    return rows


def write_excel(filepath: str, sheets: dict[str, list[dict]]):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  ⚠ openpyxl 未安装，跳过 Excel 输出")
        print("    安装: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    body_font = Font(name="Arial", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    alt_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for sheet_name, rows in sheets.items():
        if not rows:
            continue
        ws = wb.create_sheet(title=sheet_name[:31])
        headers = list(rows[0].keys())

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        for ri, row_data in enumerate(rows, 2):
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=row_data.get(h))
                cell.font = body_font
                cell.border = thin_border
                if ri % 2 == 0:
                    cell.fill = alt_fill

        for ci in range(1, len(headers) + 1):
            max_len = len(str(headers[ci - 1]))
            for row in range(2, min(len(rows) + 2, 100)):
                val = ws.cell(row=row, column=ci).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 55)

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(filepath)
    print(f"  ✅ 已保存: {filepath}")


# ─── 终端输出 ──────────────────────────────────────────────────────────

def print_summary(repos_data: dict[str, dict]):
    for repo_name, data in repos_data.items():
        runs = data.get("runs", [])
        jobs = data.get("jobs", [])
        pr_metrics = data.get("pr_metrics", [])

        conclusions = defaultdict(int)
        for j in jobs:
            conclusions[j.get("conclusion") or "unknown"] += 1
        total = len(jobs)

        job_durs = [float(j["duration_seconds"]) for j in jobs if j.get("duration_seconds") is not None]
        run_map = {r["id"]: r for r in runs}
        job_queues = [v for v in (_calc_queue_min(j, run_map.get(j["run_id"], {})) for j in jobs) if v is not None]
        # ponytail: print_summary 用秒展示，转回秒
        job_queues = [v * 60 for v in job_queues]

        print(f"\n{'='*60}")
        print(f"📊 {repo_name}")
        print(f"{'='*60}")
        print(f"  Runs:          {len(runs)}")
        print(f"  Jobs:          {total}")
        print(f"  PRs:           {len(pr_metrics)}")
        if total:
            print(f"  Job 成功率:    {conclusions.get('success', 0)}/{total} ({round(conclusions['success']/total*100, 1)}%)")
            print(f"  Job 失败率:    {conclusions.get('failure', 0)}/{total} ({round(conclusions.get('failure',0)/total*100, 1)}%)")
        if job_durs:
            avg_dur = sum(job_durs)/len(job_durs)
            print(f"  Job 平均耗时:  {avg_dur/60:.1f} 分钟")
            print(f"  Job P50 耗时:  {percentile(job_durs, 0.5)/60:.1f} 分钟")
            print(f"  Job P90 耗时:  {percentile(job_durs, 0.9)/60:.1f} 分钟")
        if job_queues:
            avg_q = sum(job_queues)/len(job_queues)
            print(f"  平均排队:      {avg_q/60:.1f} 分钟")
            print(f"  P90 排队:      {percentile(job_queues, 0.9)/60:.1f} 分钟")

        # Top 5 slowest jobs
        if job_durs:
            sorted_jobs = sorted(jobs, key=lambda j: float(j.get("duration_seconds", 0) or 0), reverse=True)[:5]
            print(f"\n  最慢的 5 个 Job:")
            for j in sorted_jobs:
                d = float(j.get("duration_seconds", 0) or 0) / 60
                print(f"    {j['name']}: {d:.1f} 分钟")

        # Top 5 workflows by run count
        wf_counts = defaultdict(int)
        for r in runs:
            wf_counts[r["name"]] += 1
        print(f"\n  最频繁的 5 个 Workflow:")
        for wf, cnt in sorted(wf_counts.items(), key=lambda x: -x[1])[:5]:
            print(f"    {wf}: {cnt} 次")


# ─── 主流程 ─────────────────────────────────────────────────────────────

def fetch_all_for_repo(client: PostgresClient, repo_id: int, date_from: str, date_to: str, skip_steps: bool = False, workflow_patterns: list[str] | None = None, workflow_files: list[str] | None = None):
    """Fetch all data for a single repo.

    传入 workflow_patterns 时，runs 在 DB 层按工作流名过滤；传入 workflow_files 时
    按 repos.yaml 文件名过滤（workflow_file/workflow_path 列）；PR 链接/指标收窄到
    只保留命中过该工作流的 PR，保证各 sheet 口径一致。
    """
    runs = fetch_runs(client, [repo_id], date_from, date_to, workflow_patterns, workflow_files)
    if not runs:
        return {"runs": [], "jobs": [], "steps": [], "pr_metrics": [], "pr_workflows": []}

    run_ids = [r["id"] for r in runs]
    jobs = fetch_jobs(client, run_ids)
    job_ids = [j["id"] for j in jobs]
    if skip_steps:
        steps = []
    else:
        steps = fetch_steps(client, job_ids)

    pr_metrics = fetch_pr_metrics(client, [repo_id], date_from, date_to)
    pr_ids = [pm["id"] for pm in pr_metrics]
    # Fix 2.3: 将 run_id 过滤下推到 SQL 层，避免 Python 端冗余过滤大量数据
    run_id_filter = set(run_ids) if workflow_patterns or workflow_files else None
    pr_workflows = fetch_pr_workflows(client, pr_ids, run_id_filter=run_id_filter)

    if workflow_patterns:
        kept_pr_ids = {pw["pr_metric_id"] for pw in pr_workflows}
        pr_metrics = [pm for pm in pr_metrics if pm["id"] in kept_pr_ids]

    return {
        "runs": runs,
        "jobs": jobs,
        "steps": steps,
        "pr_metrics": pr_metrics,
        "pr_workflows": pr_workflows,
    }


def main():
    parser = argparse.ArgumentParser(description="CI 效率分析报告生成器")
    parser.add_argument("--repo", action="append", help="仓库名 owner/repo（可多次指定）")
    parser.add_argument("--from", dest="date_from", help="起始日期 YYYY-MM-DD")
    parser.add_argument("--to", dest="date_to", help="结束日期 YYYY-MM-DD")
    parser.add_argument("--list-repos", action="store_true", help="列出所有可用仓库")
    parser.add_argument("--step-names", help="Step 分类映射 JSON 文件路径")
    parser.add_argument("--workflow", action="append", help="只统计指定工作流（名称精确匹配、不区分大小写；可多次指定）")
    parser.add_argument("--no-excel", action="store_true", help="跳过 Excel 输出")
    parser.add_argument("--skip-steps", action="store_true", help="跳过 steps 数据（加速查询）")
    parser.add_argument("--output", "-o", help="输出 Excel 文件路径（默认自动生成）")
    parser.add_argument("--config", default=str(DEFAULT_CONFIG), help="项目/workflow 对比配置（默认仓库根目录 .github-ci-efficiency.yaml）")
    parser.add_argument("--pg-url", help="PostgreSQL 连接串；默认读取 PG_DATABASE_URL")
    parser.add_argument("--success-only", action="store_true", help="job/workflow 耗时统计只算 conclusion=success 的样本（目的2 口径，ADR-005）")
    parser.add_argument("--min-duration", type=float, default=5, help="耗时下限(分钟)：低于此值的 run/job/step 不计入统计(avg/p50/p90)与关键路径，默认 5")
    parser.add_argument("--insights", action="store_true", help="额外输出 HTML 洞察报告（Top 问题+证据，ADR-005）")
    parser.add_argument("--no-drilldown", action="store_true", help="跳过下钻 HTML 报告（默认生成：>阈值分钟 run 列表 → job 条形图 → step 明细，ADR-009）")
    parser.add_argument("--drilldown-min", type=float, default=60, help="下钻报告的 run 耗时阈值(分钟)，默认 60")
    # 配置文件用于批量项目对比；--repo 可用于临时选择单个或多个仓库。
    args = parser.parse_args()

    import os
    env = load_env(ENV_FILE)
    pg_url = args.pg_url or os.getenv("PG_DATABASE_URL") or env.get("PG_DATABASE_URL") or DEFAULT_PG_URL
    client = PostgresClient(pg_url)
    all_repos = get_repo_ids(client)

    def _resolve(repo_name: str):
        if repo_name not in all_repos:
            print(f"⚠ 仓库 {repo_name} 不在本地 PostgreSQL 中")
            sys.exit(1)
        return client, all_repos[repo_name]

    # Load step names map
    step_names_path = Path(args.step_names) if args.step_names else DEFAULT_STEP_NAMES
    step_names_map = {}
    if step_names_path.exists():
        with open(step_names_path) as f:
            step_names_map = json.load(f)
        print(f"📂 加载 step 分类: {step_names_path} ({len(step_names_map)} 条映射)")

    # List repos mode
    if args.list_repos:
        print(f"\n📋 可用仓库 (共 {len(all_repos)} 个):")
        for name, rid in sorted(all_repos.items(), key=lambda x: x[1]):
            print(f"  {rid:>5}: {name}")
        return

    configured_entries = parse_config_entries(args.config) if Path(args.config).exists() else {}
    configured_workflows = {
        repo: [workflow["name"] for workflow in workflows]
        for repo, workflows in configured_entries.items()
    }
    if args.repo:
        repo_names = args.repo
    else:
        missing = [name for name in configured_workflows if name not in all_repos]
        if missing:
            print(f"⚠ 配置中的仓库尚无本地数据，跳过: {', '.join(missing)}")
        repo_names = [name for name in configured_workflows if name in all_repos]
    if not repo_names:
        print(f"❌ 配置仓库均无本地数据: {args.config}")
        sys.exit(1)

    # Date range
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    date_from = args.date_from or (datetime.now(timezone.utc) - timedelta(days=DEFAULT_DAYS)).strftime("%Y-%m-%d")
    date_to = args.date_to or today
    try:
        start = datetime.strptime(date_from, "%Y-%m-%d")
        end = datetime.strptime(date_to, "%Y-%m-%d")
    except ValueError:
        parser.error("--from/--to 必须是 YYYY-MM-DD")
    if start > end:
        parser.error("--from 不能晚于 --to")
    print(f"\n📅 时间范围: {date_from} → {date_to}")
    print(f"📦 仓库: {', '.join(repo_names)}")
    if args.workflow:
        print(f"🔍 工作流过滤: {', '.join(args.workflow)}")
    if configured_workflows and not args.workflow:
        print(f"🔍 配置 workflow 过滤: {configured_workflows}")

    # Fetch data：一个只读 PostgreSQL 连接配置服务所有仓库。
    repos_data = {}
    overview_data = []
    for repo_name in repo_names:
        client, repo_id = _resolve(repo_name)
        entries = configured_entries.get(repo_name, []) if not args.workflow else []
        wf_patterns = args.workflow or [w["name"] for w in entries if not w.get("file")] or None
        wf_files = [w["file"] for w in entries if w.get("file")] or None
        selected = args.workflow or [w["name"] for w in entries]
        print(f"\n⏳ 获取 {repo_name} (id={repo_id}) 数据..." + (f" workflow={selected}" if selected else ""))
        data = fetch_all_for_repo(
            client, repo_id, date_from, date_to,
            skip_steps=args.skip_steps,
            workflow_patterns=wf_patterns,
            workflow_files=wf_files,
        )
        normalize_configured_workflows(data["runs"], entries)
        repos_data[repo_name] = data
        print(f"  ✅ Runs: {len(data['runs'])}, Jobs: {len(data['jobs'])}, "
              f"Steps: {len(data['steps'])}, PRs: {len(data['pr_metrics'])}")
        # 总览每个 workflow 一行，避免把 NPU/GPU 聚合成一个任意命名的仓库行。
        rm = {r["id"]: r for r in data["runs"]}
        workflow_names = selected or sorted({r["name"] for r in data["runs"]})
        for workflow_name in workflow_names:
            workflow_runs = [
                r for r in data["runs"]
                if r.get("name", "").lower() == workflow_name.lower()
            ]
            workflow_run_ids = {r["id"] for r in workflow_runs}
            successful_run_ids = {
                r["id"] for r in workflow_runs if r.get("conclusion") == "success"
            }
            run_durs = [
                duration for duration in (
                    sec_to_min(r.get("duration_seconds")) for r in workflow_runs
                    if r.get("conclusion") == "success"
                ) if duration is not None and duration >= args.min_duration
            ]
            workflow_jobs = [job for job in data["jobs"] if job["run_id"] in workflow_run_ids]
            successful_jobs = [job for job in workflow_jobs if job.get("conclusion") == "success"]
            job_queues = collect_workflow_job_queues(data["jobs"], rm, workflow_run_ids)
            workflow_file = resolve_workflow_file(workflow_name, workflow_runs, entries)
            overview_data.append({
                "repo": repo_name,
                "workflow_file": workflow_file,
                "workflow_name": workflow_name,
                "total_run_count": len(workflow_runs),
                "success_run_count": len(successful_run_ids),
                "total_job_count": len(workflow_jobs),
                "success_job_count": len(successful_jobs),
                "min_duration": args.min_duration,
                "durations": run_durs,
                "queues": job_queues,
            })

    if not any(d["runs"] for d in repos_data.values()):
        print("\n⚠ 指定时间范围内没有数据")
        return

    # Terminal summary
    print_summary(repos_data)

    # Build analysis sheets
    sheets = {}

    # 总览页
    if overview_data:
        sheets["总览"] = build_overview(overview_data)

    # Multi-repo comparison
    if len(repo_names) > 1:
        sheets["仓库对比"] = analyze_comparison(repos_data)

    # Per-repo analysis
    for repo_name, data in repos_data.items():
        runs, jobs, steps = data["runs"], data["jobs"], data["steps"]
        pr_metrics, pr_workflows = data["pr_metrics"], data["pr_workflows"]

        sheet_prefix = repo_name if len(repo_names) > 1 else None

        if len(runs) == 0:
            continue

        wf = analyze_workflow_stats(runs, jobs, success_only=args.success_only, min_duration=args.min_duration)
        sheets[_sp(sheet_prefix, "工作流统计")] = wf

        js = analyze_job_stats(runs, jobs, success_only=args.success_only, min_duration=args.min_duration)
        sheets[_sp(sheet_prefix, "任务统计")] = js

        if steps:
            ss = analyze_step_stats(steps, step_names_map, success_only=args.success_only, min_duration=args.min_duration)
            sheets[_sp(sheet_prefix, "步骤统计")] = ss

        ps = analyze_pr_stats(pr_metrics, pr_workflows)
        sheets[_sp(sheet_prefix, "PR统计")] = ps

        pd_rows = build_pr_details(pr_metrics, pr_workflows, runs, jobs, steps)
        sheets[_sp(sheet_prefix, "PR详情")] = pd_rows

    # Write Excel
    if not args.no_excel and sheets:
        if args.output:
            outfile = args.output
        else:
            date_tag = f"{date_from}_to_{date_to}"
            repo_tag = "_vs_".join(r.replace("/", "_") for r in repo_names)
            outfile = f"{repo_tag}-ci-report-{date_tag}.xlsx"
        write_excel(outfile, sheets)
        print(f"\n📊 共生成 {len(sheets)} 个 sheet: {', '.join(sheets.keys())}")

    # HTML 洞察报告（--insights，ADR-005）
    if args.insights:
        for repo_name, data in repos_data.items():
            runs, jobs, steps = data["runs"], data["jobs"], data["steps"]
            html_out = outfile.replace(".xlsx", "-insights.html") if len(repo_names) == 1 else f"{repo_name.replace('/', '_')}-insights-{date_from}_to_{date_to}.html"
            api_info = "数据源：Action Insight 本地 PostgreSQL"
            write_html_report(html_out, repo_name, date_from, date_to, runs, jobs, steps,
                              step_names_map, args.success_only, args.min_duration, api_info)

    # 下钻 HTML（默认生成，ADR-009）
    if not args.no_drilldown:
        date_tag = f"{date_from}_to_{date_to}"
        repo_tag = "_vs_".join(r.replace("/", "_") for r in repo_names)
        drill_out = (args.output.replace(".xlsx", "-drilldown.html") if args.output
                     else f"{repo_tag}-drilldown-{date_tag}.html")
        api_info = "数据源：Action Insight 本地 PostgreSQL"
        write_drilldown_html(drill_out, repos_data, date_from, date_to, step_names_map, api_info, min_minutes=args.drilldown_min)


if __name__ == "__main__":
    main()