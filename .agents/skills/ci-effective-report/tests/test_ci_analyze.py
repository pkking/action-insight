import importlib.util
import sys
import unittest
from pathlib import Path

SCRIPT = Path(__file__).parents[1] / "scripts" / "ci_analyze.py"
SPEC = importlib.util.spec_from_file_location("ci_analyze", SCRIPT)
MODULE = importlib.util.module_from_spec(SPEC)
assert SPEC.loader
sys.modules["ci_analyze"] = MODULE
SPEC.loader.exec_module(MODULE)

DUR_MIN = 60  # 测试阈值与默认一致


def _run(rid, name, dur_sec, event="push", created="2026-07-15T10:00:00Z", conclusion="success"):
    return {
        "id": rid, "repo_id": 1, "name": name, "head_branch": "main", "head_sha": "abc",
        "event": event, "status": "completed", "conclusion": conclusion,
        "created_at": created, "updated_at": "2026-07-15T11:30:00Z",
        "html_url": f"https://github.com/o/r/actions/runs/{rid}",
        "duration_seconds": dur_sec, "date": "2026-07-15", "workflow_file": "e2e.yml",
    }


def _job(jid, run_id, name, dur_sec, started="2026-07-15T10:05:00Z"):
    return {
        "id": jid, "run_id": run_id, "name": name, "status": "completed", "conclusion": "success",
        "created_at": "2026-07-15T10:00:00Z", "started_at": started, "completed_at": "2026-07-15T11:00:00Z",
        "html_url": f"https://github.com/o/r/jobs/{jid}", "queue_duration_seconds": 300, "duration_seconds": dur_sec,
    }


def _step(job_id, num, name, dur_sec, conclusion="success"):
    return {
        "job_id": job_id, "number": num, "name": name, "status": "completed", "conclusion": conclusion,
        "started_at": "2026-07-15T10:10:00Z", "completed_at": "2026-07-15T10:40:00Z", "duration_seconds": dur_sec,
    }


class ConfigTests(unittest.TestCase):
    def test_parse_config_reads_all_workflows_per_repo(self):
        config = Path("/tmp/ci-effective-config.yaml")
        config.write_text("""repositories:\n  - repo: o/r\n    workflows:\n      - name: E2E\n        file: e2e.yml\n      - name: Nightly\n""")
        self.assertEqual(MODULE.parse_config(str(config)), {"o/r": ["E2E", "Nightly"]})
        self.assertEqual(MODULE.parse_config_entries(str(config))["o/r"][0]["file"], "e2e.yml")

    def test_workflow_file_normalizes_dynamic_run_name(self):
        runs = [{"name": "PR #123 - dynamic title", "workflow_file": "pr-test.yml"}]
        MODULE.normalize_configured_workflows(
            runs,
            [{"name": "PR Test Base", "file": "pr-test.yml"}],
        )
        self.assertEqual(runs[0]["name"], "PR Test Base")
        self.assertEqual(runs[0]["run_name"], "PR #123 - dynamic title")

    def test_workflow_file_falls_back_to_most_common_run_metadata(self):
        runs = [
            {"workflow_file": "ci.yml"},
            {"workflow_file": "ci.yml"},
            {"workflow_file": "old.yml"},
        ]
        self.assertEqual(MODULE.resolve_workflow_file("CI", runs, [{"name": "CI"}]), "ci.yml")

    def test_queue_samples_include_successful_jobs_from_failed_workflows(self):
        runs = {1: _run(1, "Nightly", 60, conclusion="failure")}
        jobs = [_job(10, 1, "setup", 30, started="2026-07-15T10:05:00Z")]
        self.assertEqual(MODULE.collect_workflow_job_queues(jobs, runs, {1}), [5.0])

    def test_queue_excludes_upstream_job_execution_time(self):
        run = _run(1, "E2E", 14_400)
        job = _job(10, 1, "dependent test", 300, started="2026-07-15T14:05:00Z")
        job["created_at"] = "2026-07-15T14:00:00Z"
        self.assertEqual(MODULE._calc_queue_min(job, run), 5.0)

    def test_workflow_resource_summary_groups_models_and_uses_maximum_single_run(self):
        jobs = [
            {"run_id": 1, "card_model": "linux-aarch64-a3", "card_count": 4},
            {"run_id": 1, "card_model": "linux-aarch64-a2b3", "card_count": 1},
            {"run_id": 1, "card_model": "linux-aarch64-a2", "card_count": 2},
            {"run_id": 1, "labels": ["linux-amd64-xx-cpu-4"]},
            {"run_id": 2, "card_model": "linux-aarch64-a3", "card_count": 2},
        ]
        self.assertEqual(MODULE.workflow_resource_summary(jobs), "A2 × 3卡；A3 × 2卡；x86 CPU × 4核")

    def test_workflow_resource_summary_prefers_static_workflow_requirements(self):
        jobs = [{"run_id": 1, "card_model": "linux-aarch64-a3", "card_count": 64}]
        static = {"L20": 8}
        self.assertEqual(MODULE.workflow_resource_summary(jobs, static), "L20 × 8卡")

    def test_overview_distinguishes_total_runs_from_valid_successes(self):
        row = MODULE.build_overview([{
            "repo": "o/r",
            "workflow_name": "GPU",
            "total_run_count": 54,
            "success_run_count": 0,
            "total_job_count": 16,
            "success_job_count": 0,
            "min_duration": 5,
            "durations": [],
            "queues": [],
        }])[0]
        self.assertEqual(row["总Run数"], 54)
        self.assertEqual(row["成功Run数"], 0)
        self.assertEqual(row["有效成功Run数"], 0)
        self.assertIsNone(row["E2E 平均(分钟)"])
        self.assertEqual(
            row["空值判断依据"],
            "总 Run=54，成功 Run=0，E2E 不计算；Jobs=16，成功 Jobs=0，排队不计算",
        )

    def test_resource_pool_usage_is_clipped_to_window_and_hour(self):
        job = _job(1, 1, "a3", 7_200, started="2026-07-15T00:30:00Z")
        job.update({"completed_at": "2026-07-15T02:30:00Z", "card_model": "linux-aarch64-a3", "card_count": 4})
        summary, timeline = MODULE.build_resource_pool_rows({"o/r": {"jobs": [job]}}, "2026-07-15", "2026-07-15", {"A3": 10})
        self.assertEqual(summary[0]["消耗卡时"], 4.0)
        self.assertEqual(summary[0]["时间校正后总卡时"], 20.0)
        self.assertEqual([row["消耗卡时"] for row in timeline], [1.0, 2.0, 1.0])

    def test_overview_explains_workflow_with_no_runs(self):
        reason = MODULE.overview_missing_reason(
            {"total_run_count": 0}, [], [],
        )
        self.assertEqual(reason, "窗口内无 Run，E2E 与排队均无法计算")

    def test_workflow_filter_is_exact_and_case_insensitive(self):
        clause = MODULE._workflow_match_clause(["E2E"])
        self.assertIn("ILIKE 'E2E'", clause)
        self.assertNotIn("%E2E%", clause)

    def test_report_repositories_are_collected_by_etl(self):
        import yaml
        root = Path(__file__).parents[4]
        report_repos = set(MODULE.parse_config(str(root / ".github-ci-efficiency.yaml")))
        etl = yaml.safe_load((root / "etl" / "repos.yaml").read_text())
        etl_repos = {item["repo"] for item in etl["repos"]}
        self.assertEqual(report_repos - etl_repos, set())

    def test_excel_uses_arial(self):
        from openpyxl import load_workbook
        output = "/tmp/ci-effective-font.xlsx"
        MODULE.write_excel(output, {"Summary": [{"repo": "o/r"}]})
        ws = load_workbook(output)["Summary"]
        self.assertEqual(ws["A1"].font.name, "Arial")
        self.assertEqual(ws["A2"].font.name, "Arial")

    def test_overview_headers_have_metric_comments(self):
        from openpyxl import load_workbook
        output = "/tmp/ci-effective-overview.xlsx"
        MODULE.write_excel(output, {"总览": [{"仓库": "o/r", "E2E P90(分钟)": 10}]})
        ws = load_workbook(output)["总览"]
        self.assertIn("P90", ws["B1"].comment.text)

    def test_resource_pool_sheet_has_pie_chart(self):
        from openpyxl import load_workbook
        output = "/tmp/ci-effective-resource-pool.xlsx"
        MODULE.write_excel(output, {"资源池利用率": [{"项目": "o/r", "资源类型": "A3", "资源池卡数": 10, "消耗卡时": 5}]})
        self.assertEqual(len(load_workbook(output)["资源池利用率"]._charts), 1)


class FetchJobsTests(unittest.TestCase):
    def test_recent_resource_hints_use_latest_job_labels(self):
        class Client:
            def query(self, _sql):
                return [{"id": 1, "run_id": 2, "labels_json": '["linux-aarch64-a2-1"]', "card_model": None, "card_count": None}]

        hints = MODULE.fetch_recent_resource_jobs(Client(), 1, [{"name": "CI", "file": "ci.yml"}])
        self.assertEqual(MODULE.workflow_resource_summary(hints["CI"]), "A2 × 1卡")

    def test_fetch_jobs_falls_back_to_attempt_scoped_storage(self):
        class Client:
            sql = ""

            def query(self, sql):
                self.sql = sql
                return []

        client = Client()
        MODULE.fetch_jobs(client, [1])
        self.assertIn("FROM workflow_jobs", client.sql)
        self.assertIn("NOT EXISTS (SELECT 1 FROM jobs j WHERE j.id = wj.job_id)", client.sql)


class BuildDrilldownDataTests(unittest.TestCase):
    def _repos_data(self):
        # run1: 90min keep; run2: 40min drop; run3: 120min keep (schedule, no PR author)
        runs = [_run(1, "E2E <prod>", 90 * 60), _run(2, "Short", 40 * 60),
                _run(3, "Nightly & CI", 120 * 60, event="schedule")]
        jobs = [_job(10, 1, "build", 70 * 60), _job(11, 1, "test", 30 * 60),
                _job(12, 3, "nightly-1", 110 * 60)]
        # step number 2 sorts after number 1; None-number sorts last
        steps = [_step(10, 2, "Run <tests>", 65 * 60), _step(10, 1, "checkout", 5 * 60),
                 _step(12, None, "no-number step", 100 * 60)]
        pr_metrics = [{"id": 99, "author": "alice", "html_url": "https://github.com/o/r/pull/1",
                       "pr_number": 1, "title": "t", "created_at": "2026-07-15T09:00:00Z",
                       "merged_at": "2026-07-15T12:00:00Z", "ci_completed_at": "2026-07-15T11:30:00Z",
                       "conclusion": "success"}]
        pr_workflows = [{"pr_metric_id": 99, "run_id": 1}]
        return {"o/r": {"runs": runs, "jobs": jobs, "steps": steps,
                        "pr_metrics": pr_metrics, "pr_workflows": pr_workflows}}

    def test_threshold_excludes_short_runs(self):
        data = MODULE.build_drilldown_data(self._repos_data(), None, min_minutes=DUR_MIN)
        self.assertEqual([r["dur"] for r in data["runs"]], [120.0, 90.0])

    def test_runs_sorted_desc_by_duration(self):
        data = MODULE.build_drilldown_data(self._repos_data(), None, min_minutes=DUR_MIN)
        self.assertEqual(data["runs"][0]["dur"], 120.0)
        self.assertEqual(data["runs"][1]["dur"], 90.0)

    def test_author_joined_via_pr_workflows_and_empty_for_non_pr(self):
        data = MODULE.build_drilldown_data(self._repos_data(), None, min_minutes=DUR_MIN)
        by_id = {r["url"]: r for r in data["runs"]}
        self.assertEqual(by_id["https://github.com/o/r/actions/runs/1"]["author"], "alice")
        self.assertEqual(by_id["https://github.com/o/r/actions/runs/3"]["author"], "")

    def test_jobs_chronological_and_steps_sorted_by_number_none_last(self):
        data = MODULE.build_drilldown_data(self._repos_data(), None, min_minutes=DUR_MIN)
        run1 = next(r for r in data["runs"] if r["dur"] == 90.0)
        # jobs sorted by started_at ascending (build 10:05 before test 10:30)
        self.assertEqual([j["name"] for j in run1["jobs"]], ["build", "test"])
        # job timestamps are carried through for the Gantt timeline
        self.assertEqual(run1["jobs"][0]["started"], "2026-07-15T10:05:00Z")
        self.assertEqual(run1["jobs"][0]["created"], "2026-07-15T10:00:00Z")
        self.assertEqual(run1["jobs"][0]["completed"], "2026-07-15T11:00:00Z")
        steps = run1["jobs"][0]["steps"]
        self.assertEqual([s["n"] for s in steps], [1, 2])

    def test_step_type_classification(self):
        data = MODULE.build_drilldown_data(self._repos_data(), None, min_minutes=DUR_MIN)
        run1 = next(r for r in data["runs"] if r["dur"] == 90.0)
        types = {s["name"]: s["type"] for s in run1["jobs"][0]["steps"]}
        self.assertEqual(types["checkout"], "构建")
        self.assertEqual(types["Run <tests>"], "执行测试")

    def test_missing_duration_renders_as_none_not_zero(self):
        runs = [_run(1, "E2E", 90 * 60)]
        jobs = [_job(10, 1, "build", 70 * 60)]
        steps = [_step(10, 1, "broken", None)]
        repos = {"o/r": {"runs": runs, "jobs": jobs, "steps": steps, "pr_metrics": [], "pr_workflows": []}}
        data = MODULE.build_drilldown_data(repos, None, min_minutes=DUR_MIN)
        self.assertIsNone(data["runs"][0]["jobs"][0]["steps"][0]["dur"])

    def test_stats_based_on_valid_runs_over_10min(self):
        # runs: 5min(无效<10), 25min(有效但<60), 90min(有效且>60); 显示阈值=60
        # job 排队：run1 的 job 排队 5min、run2 的 job 排队 20min、run3 的 job 排队 1min
        runs = [_run(1, "a", 5 * 60), _run(2, "b", 25 * 60), _run(3, "c", 90 * 60)]
        jobs = [
            {"id": 1, "run_id": 2, "name": "j", "status": "completed", "conclusion": "success",
             "created_at": "2026-07-15T10:00:00Z", "started_at": "2026-07-15T10:20:00Z",
             "completed_at": "2026-07-15T10:25:00Z", "html_url": "", "queue_duration_seconds": 1200, "duration_seconds": 300},
            {"id": 2, "run_id": 3, "name": "j", "status": "completed", "conclusion": "success",
             "created_at": "2026-07-15T10:00:00Z", "started_at": "2026-07-15T10:01:00Z",
             "completed_at": "2026-07-15T11:30:00Z", "html_url": "", "queue_duration_seconds": 60, "duration_seconds": 5340},
        ]
        repos = {"o/r": {"runs": runs, "jobs": jobs, "steps": [], "pr_metrics": [], "pr_workflows": []}}
        data = MODULE.build_drilldown_data(repos, None, min_minutes=DUR_MIN)
        # 表格只入 >60 的
        self.assertEqual([r["dur"] for r in data["runs"]], [90.0])
        # 统计基于 >10（有效）
        s = data["stats"]["o/r"]
        self.assertEqual(s["valid"], 2)   # 25 + 90
        self.assertEqual(s["over60"], 1) # 90
        self.assertNotIn("avg", s)
        self.assertNotIn("q_avg", s)
        self.assertNotIn("card_hours", s)
        self.assertNotIn("npu_p50", s)
        self.assertNotIn("npu_pass_rate", s)
        # P50/P90 and queue stats are run-level
        self.assertIn("p50", s)
        self.assertIn("q_p50", s)
        self.assertIn("pass_rate", s)
        self.assertEqual(s["npu_hours"], 0)

    def test_all_runs_includes_every_run_not_just_threshold_runs(self):
        # 3 runs: 5min, 25min, 90min; display threshold=60 -> table shows only 90min
        runs = [_run(1, "a", 5 * 60), _run(2, "b", 25 * 60), _run(3, "c", 90 * 60)]
        jobs = [_job(10, 3, "j", 80 * 60)]
        repos = {"o/r": {"runs": runs, "jobs": jobs, "steps": [], "pr_metrics": [], "pr_workflows": []}}
        data = MODULE.build_drilldown_data(repos, None, min_minutes=DUR_MIN)
        # table only shows >60
        self.assertEqual(len(data["runs"]), 1)
        # all_runs has all 3
        self.assertEqual(len(data["all_runs"]), 3)
        self.assertEqual({r["wf"] for r in data["all_runs"]}, {"a", "b", "c"})
        # all_runs entries have run-level fields, no nested jobs/steps
        r0 = data["all_runs"][0]
        for key in ("repo", "author", "created", "updated", "wf", "event", "dur",
                    "card_hours", "cpu_hours", "status", "conclusion", "url"):
            self.assertIn(key, r0)
        self.assertNotIn("unknown_card_jobs", r0)

    def test_card_hours_cover_all_runs_independent_of_table_threshold(self):
        runs = [
            _run(1, "long", 90 * 60),
            _run(2, "short failure", 5 * 60, conclusion="failure"),
        ]
        known = _job(10, 1, "eight-card", 55 * 60)
        known["card_count"] = 8
        known["card_model"] = "310p"
        failed = _job(11, 2, "two-card", 55 * 60)
        failed["card_count"] = 2
        failed["card_model"] = "a3"
        unknown = _job(12, 1, "unknown", 55 * 60)
        never_started = _job(13, 1, "not-started", 55 * 60)
        never_started["card_count"] = 4
        never_started["started_at"] = ""
        cpu = _job(14, 1, "pre-commit", 10 * 60)
        cpu["labels"] = ["linux-amd64-cpu-8-hk"]
        repos = {"o/r": {"runs": runs, "jobs": [known, failed, unknown, never_started, cpu],
                          "steps": [], "pr_metrics": [], "pr_workflows": []}}

        data = MODULE.build_drilldown_data(repos, None, min_minutes=DUR_MIN)
        self.assertEqual([r["wf"] for r in data["runs"]], ["long"])
        # NPU card-hours: 8×55min + 1×55min (A3 has two dies/card) = 495min; CPU hours: 10min
        stats = data["stats"]["o/r"]
        self.assertAlmostEqual(stats["npu_hours"], 495 / 60, places=5)
        self.assertAlmostEqual(stats["npu_failure_hours"], 55 / 60, places=5)
        self.assertAlmostEqual(stats["npu_by_resource"]["310P"], 440 / 60, places=5)
        self.assertAlmostEqual(stats["npu_by_resource"]["A3"], 55 / 60, places=5)
        self.assertAlmostEqual(stats["cpu_hours"], 55 / 60, places=5)
        self.assertNotIn("unknown_card_jobs", stats)
        self.assertNotIn("avg", stats)
        self.assertNotIn("q_avg", stats)
        self.assertNotIn("card_hours", stats)
        long_run = data["runs"][0]
        self.assertAlmostEqual(long_run["card_hours"], 440 / 60, places=5)
        self.assertAlmostEqual(long_run["cpu_hours"], 55 / 60, places=5)
        job_hours = {job["name"]: job["card_hours"] for job in long_run["jobs"]}
        self.assertAlmostEqual(job_hours["eight-card"], 440 / 60, places=5)
        self.assertIsNone(job_hours["unknown"])

    def test_run_with_no_jobs_and_failed_cancelled_conclusions(self):
        # run with zero jobs, plus a run whose job failed and another cancelled
        runs = [_run(1, "no-jobs", 90 * 60), _run(2, "mixed", 90 * 60, conclusion="failure")]
        jobs = [_job(20, 2, "failed-job", 70 * 60)]
        jobs[0]["conclusion"] = "failure"
        cancelled = _job(21, 2, "cancelled-job", 5 * 60)
        cancelled["conclusion"] = "cancelled"
        jobs.append(cancelled)
        steps = [_step(20, 1, "s", 65 * 60, conclusion="failure")]
        repos = {"o/r": {"runs": runs, "jobs": jobs, "steps": steps, "pr_metrics": [], "pr_workflows": []}}
        data = MODULE.build_drilldown_data(repos, None, min_minutes=DUR_MIN)
        nojobs = next(r for r in data["runs"] if r["wf"] == "no-jobs")
        self.assertEqual(nojobs["jobs"], [])  # run with no jobs → empty list, renderJobs shows "无 job 数据"
        mixed = next(r for r in data["runs"] if r["wf"] == "mixed")
        conclusions = {j["name"]: j["conclusion"] for j in mixed["jobs"]}
        self.assertEqual(conclusions["failed-job"], "failure")
        self.assertEqual(conclusions["cancelled-job"], "cancelled")
        self.assertEqual(mixed["conclusion"], "failure")  # run-level failure preserved



class WriteDrilldownHtmlTests(unittest.TestCase):
    def test_html_escapes_names_and_has_no_premature_script_close(self):
        runs = [_run(1, "E2E </script><img>", 90 * 60)]
        jobs = [_job(10, 1, "job</script>x", 70 * 60)]
        steps = [_step(10, 1, "step </script> y", 65 * 60)]
        repos = {"o/r": {"runs": runs, "jobs": jobs, "steps": steps, "pr_metrics": [], "pr_workflows": []}}
        out = "/tmp/test-drilldown.html"
        MODULE.write_drilldown_html(out, repos, "2026-07-01", "2026-07-31", {}, "test", min_minutes=DUR_MIN)
        html = Path(out).read_text(encoding="utf-8")
        blob = html.split("const DATA=", 1)[1].split(";\nfunction esc", 1)[0]
        # our guard turns </ into <\/ so a malicious </script> never closes the tag early
        self.assertNotIn("</script>", blob)
        # the table headers are present and server-rendered
        self.assertIn("代码仓", html)
        self.assertIn("提交人", html)
        self.assertIn("结束时间", html)
        self.assertIn("Run URL", html)
        self.assertIn("NPU", html)
        self.assertIn("CPU", html)
        self.assertIn("失败机时", html)
        self.assertIn("资源卡时", html)
        self.assertIn("达标率", html)
        self.assertNotIn("未知卡数 Job", html)
        self.assertNotIn("平均耗时", html)
        self.assertNotIn("平均排队", html)
        # round-trip: undo the guard and the JSON is valid
        import json
        parsed = json.loads(blob.replace("<\\/", "</"))
        self.assertEqual(parsed["runs"][0]["wf"], "E2E </script><img>")
        self.assertEqual(parsed["runs"][0]["jobs"][0]["name"], "job</script>x")

    def test_empty_repos_data_produces_valid_html(self):
        MODULE.write_drilldown_html("/tmp/test-drilldown-empty.html", {}, "2026-07-01", "2026-07-31", {}, "t", min_minutes=DUR_MIN)
        html = Path("/tmp/test-drilldown-empty.html").read_text(encoding="utf-8")
        self.assertIn("命中 <b>0</b> 个 run", html)

    def test_title_and_tab_structure_per_repo(self):
        # two repos -> two tabs, each with a per-project header, no "下钻" wording in title
        repos = {"o/r": {"runs": [_run(1, "E2E", 90 * 60)], "jobs": [], "steps": [], "pr_metrics": [], "pr_workflows": []},
                 "a/b": {"runs": [_run(2, "CI", 80 * 60)], "jobs": [], "steps": [], "pr_metrics": [], "pr_workflows": []}}
        MODULE.write_drilldown_html("/tmp/test-drilldown-tabs.html", repos, "2026-07-01", "2026-07-31", {}, "t", min_minutes=DUR_MIN)
        html = Path("/tmp/test-drilldown-tabs.html").read_text(encoding="utf-8")
        self.assertIn("<h1>CI 效率报告</h1>", html)
        self.assertNotIn("下钻报告", html)
        # per-project header is built by JS at runtime; assert the template fragment, not a static literal
        self.assertIn(" CI效率报告</h2>", html)
        self.assertIn('class="tabs"', html)
        # JS groups by repo and renders per-repo panels
        self.assertIn("const REPOS=", html)
        self.assertIn("BY_REPO=", html)
        self.assertIn("const ri=activeRepo", html)
        self.assertIn("renderRows(ri);", html)

    def test_gantt_timeline_has_queue_and_run_bars(self):
        # renderJobs must emit a shared-axis timeline with orange queue + blue run bars
        runs = [_run(1, "E2E", 90 * 60)]
        jobs = [_job(10, 1, "build", 70 * 60)]
        steps = [_step(10, 1, "checkout", 5 * 60)]
        repos = {"o/r": {"runs": runs, "jobs": jobs, "steps": steps, "pr_metrics": [], "pr_workflows": []}}
        MODULE.write_drilldown_html("/tmp/test-drilldown-gantt.html", repos, "2026-07-01", "2026-07-31", {}, "t", min_minutes=DUR_MIN)
        html = Path("/tmp/test-drilldown-gantt.html").read_text(encoding="utf-8")
        self.assertIn('class="gantt"', html)
        self.assertIn('class="gantt-ruler"', html)
        self.assertIn('class="bar queue"', html)   # 橙色排队段
        self.assertIn('class="bar run"', html)     # 蓝色运行段
        self.assertIn('class="gantt-track"', html)
        self.assertIn(".steps { overflow-x: auto; padding", html)
        self.assertIn("overflow-wrap: anywhere", html)
        self.assertIn("grid-template-columns: 180px minmax(120px, 1fr) 220px", html)
        self.assertIn(".gantt { background: #fff; min-width: 1260px", html)
        self.assertIn("fmtT(", html)  # axis timestamp formatting

    def test_timing_causes_in_run_detail(self):
        runs = [_run(1, "E2E", 90 * 60)]
        # job1: 5min queue, 70min execution; job2: 2min queue, 30min execution
        j1 = _job(10, 1, "long-job", 70 * 60, started="2026-07-15T10:05:00Z")
        j2 = _job(11, 1, "short-job", 30 * 60, started="2026-07-15T10:02:00Z")
        steps = [_step(10, 1, "checkout", 5 * 60), _step(10, 2, "build", 60 * 60)]
        repos = {"o/r": {"runs": runs, "jobs": [j1, j2], "steps": steps, "pr_metrics": [], "pr_workflows": []}}
        data = MODULE.build_drilldown_data(repos, None, min_minutes=DUR_MIN)
        r = data["runs"][0]
        self.assertIn("timing_causes", r)
        tc = r["timing_causes"]
        self.assertGreater(len(tc), 0)
        # longest job execution should be present
        kinds = [c["kind"] for c in tc]
        self.assertIn("Job execution", kinds)
        # should be sorted by duration desc
        self.assertEqual(tc, sorted(tc, key=lambda x: -x["dur"]))

    def test_csv_export_button_and_function_present(self):
        repos = {"o/r": {"runs": [_run(1, "E2E", 90 * 60)], "jobs": [], "steps": [], "pr_metrics": [], "pr_workflows": []}}
        MODULE.write_drilldown_html("/tmp/test-drilldown-csv.html", repos, "2026-07-01", "2026-07-31", {}, "t", min_minutes=DUR_MIN)
        html = Path("/tmp/test-drilldown-csv.html").read_text(encoding="utf-8")
        self.assertIn("导出 CSV", html)
        self.assertIn("exportCSV(", html)

    def test_toggle_uses_table_row_not_empty_string(self):
        # regression: setting display='' falls back to CSS display:none, hiding the row forever
        runs = [_run(1, "E2E", 90 * 60)]
        jobs = [_job(10, 1, "build", 70 * 60)]
        steps = [_step(10, 1, "checkout", 5 * 60)]
        repos = {"o/r": {"runs": runs, "jobs": jobs, "steps": steps, "pr_metrics": [], "pr_workflows": []}}
        MODULE.write_drilldown_html("/tmp/test-drilldown-toggle.html", repos, "2026-07-01", "2026-07-31", {}, "t", min_minutes=DUR_MIN)
        html = Path("/tmp/test-drilldown-toggle.html").read_text(encoding="utf-8")
        # showing must set an explicit display that beats the .detail{display:none} rule
        self.assertIn("table-row", html)
        self.assertIn("const open=det.style.display==='table-row'", html)


if __name__ == "__main__":
    unittest.main()
